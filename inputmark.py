#! python 3
# _*_ coding: utf_8  _*_

import openpyxl
import os
import re
import logging

# logging.basicConfig( level = logging.ERROR, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )

class Writemarks():
    """ help input the mark. just input the number of the student, then the mark.
    """

    def __init__(self):
        self.marklist = []
        pass

    def inputmarks(self, dstdir):
        finish = ''
        while not finish:
            studnum = input("input a student's number:")
            mark = input("input the student's mark:")
            self.marklist.append((studnum, mark))
            finish = input("input any letter to end. Return directly means continue.")
            
        wb = openpyxl.Workbook()
        sheet = wb.get_active_sheet()
        for row in range(len(self.marklist)):
            sheet.cell(row=row + 2,column = 2).value = self.marklist[row][0]
            sheet.cell(row=row + 2,column = 3).value = self.marklist[row][1]
        dstbook = dstdir + '\\' + 'BKmark-1520603-4' + ".xlsx"
        wb.save(dstbook)
        



def main():
    SRCDIR = 'd:\\_PythonWorks\\excelOperate\\cj-2016201701'
    DSTDIR = 'd:\\_PythonWorks\\excelOperate\\bkcj-201702'
    writem = Writemarks()

    # input the students' mark to form a list.
    writem.inputmarks(DSTDIR)
    input('input any letter to finish.')
    
    # open workbook that include the marks of the students.
    FILENAME = input("Please input the file name that you will read.")
    wb = openpyxl.load_workbook(FILENAME)

##    """ find students that have failed to pass the course,
##    and copy the performance mark of the students.
##    """
##    if( int(sheet['J'+str(row)].value) < 59 ):
##        
##    # open the new workbook.
##    
##    # write the performance mark and reexam's mark. and calculate the total mark.
##


if __name__ == '__main__':
    logging.critical('--------Start of program---------')
    main()
    logging.critical('--------End---------')

##
##
##
##    CLASSRE = '-(\d{7}z?)\.' # CLASSREG = re.compile(r'\d{7}[zZ]?')
##    STUDRE = '\d{12}z?'  # STUDREG = re.compile(r'\d{12}z?')
##    STUDNUM = input("please input student's number: ")
##
##    collmark = collectmarks()
##    CLASSLIST = collmark.classfind(SRCDIR,RELIST)    
##    collmark.copymark(SRCDIR)    
##    collmark.writeclass(DSTDIR)
##
##
##
### find the Chinese words '学生名单' in filename
##ChineseReg = re.compile(r'学生名单')
### find the class
##classReg = re.compile(r'\d{7}')
### find the course
##courseReg = re.compile(r'-(\w{3,11})-')
##
##performanceTag = [
##    ['旷课',],
##    ['迟到',],
##    ['早退',],
##    ['提出问题',],
##    ['回答问题',],
##    ]
##labTag = [
##    ['旷课',],
##    ['迟到',],
##    ['早退',],
##    ['操作1',],['操作2',],['操作3',],['操作4',],['操作5',],['操作6',],['操作7',],['操作8',],
##    ['数据1',],['数据2',],['数据3',],['数据4',],['数据5',],['数据6',],['数据7',],['数据8',],   
##    ]
##designTag = [
##    ['旷课',],
##    ['迟到',],
##    ['早退',],
##    ['设计1',],['设计2',],['设计3',],['设计4',],
##    ['数据1',],['数据2',],['数据3',],['数据4',],  
##    ]
##practiceTag = [
##    ['旷课',],
##    ['迟到',],
##    ['早退',],
##    ['操作1',],['操作2',],['操作3',],['操作4',],
##    ['数据1',],['数据2',],['数据3',],['数据4',],   
##    ]
##
### 提示输入信息
##filelist = []
##fulllist = []
##k = 0
##dirname = 'd:\\_PythonWorks\\excelOperate\\pscj161702'
##for file in os.listdir(dirname):
##    fullname = dirname + '\\' + file
##    if not ChineseReg.search(file):    # ChineseReg = re.compile(r'学生名单')
##        filelist.append(file)
##        fulllist.append(fullname)     
##        print(k,file)
##        k += 1      
##tagdict = {'performance':performanceTag, 'lab':labTag, 'design':designTag, 'practice':practiceTag}
##courseNum = int(input('\n please input a number for 课程: '))    
##coursetype = courseReg.search(filelist[courseNum]).group(1)
##   
##logging.critical(fulllist[courseNum])
##wb = openpyxl.load_workbook(fulllist[courseNum])
##sheet = wb.get_active_sheet()
##
##finish = 0
##while not finish:  
##    print(coursetype)
##    item = {}
##    k = 6
##    for val in tagdict[coursetype]:
##        print(str(k) + ' ' + str(val) + ' ')
##        item[k] = str(val)
##        k += 1
##
##    itemNum = int(input('\n please input a numbers for select item: '))
##    stuNum = input("\n please input three last digitals of select student's number: 205 ")
##    mark = input('\n please input the mark: ')
##    
##    for row in range(3,sheet.max_row):
##        logging.debug(str(sheet['b'+str(row)].value)[-3:])           #  学号在B列
##        if str(sheet['b'+str(row)].value)[-3:] == stuNum:                   #  学号在B列
##            # Write
##            logging.critical(sheet.cell(row = row,column = itemNum).value)   # 写之前，cell的值
##            sheet.cell(row = row,column = itemNum).value += int(mark)        # 加上要加减的分数
##            sheet.cell(row = row,column = 4).value += int(mark)              # 总分也加上该分数
##            logging.critical(sheet.cell(row = row,column = itemNum).value)   # 写之后，cell的值
##            print(stuNum,' 总分是：', sheet.cell(row = row,column = 4).value)
##            break
##    finish = input('input any letter to finish.')
##marks = []
##for row in range(3,sheet.max_row):
##    marks.append((sheet.cell(row = row,column = 4).value, sheet['b'+str(row)].value, sheet['c'+str(row)].value))
##wb.save(fulllist[courseNum])
##marks.sort(reverse=True)
##print('前5名为：')
##for k in marks[:5]:
##    print(k)
##logging.critical('---------------')
##print('后5名为：')
##for k in marks[-5:]:
##    print(k)
##
##logging.critical('-------End--------')
##
##
##
##
##
##
##
