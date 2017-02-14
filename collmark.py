# _*_ coding: utf_8  _*_
# copy marks in many workbooks and put them togather for every student.

import openpyxl
import os
import re
import logging

logging.basicConfig( level = logging.ERROR, format = ' %(asctime)s - %(levelname)s - %(message)s' )
# logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )

class collectmarks():
    """    collect marks of the students in several courses.    
    """

    def __init__(self):
        self.classlist = []
        self.coursemark = []
        self.allcoursemark = {}
        self.studentmarks = []

    def classfind(self, srcdir, relist):
        for file in os.listdir(srcdir):
            logging.info(file)
            CLASSREG = re.compile(relist)
            if CLASSREG.search(file):
                # tuple as a element of the list
                self.classlist.append((CLASSREG.search(file).group(1),file))     
        return self.classlist

    def copymark(self, srcdir):
        self.allcoursemark = {}    # for store marks of all courses, class as key, marks as the values    
        for t in self.classlist:
            self.coursemark = []   # for store marks of one course
            classname = t[0]
            file = t[1]            
            fullname = srcdir + '\\' + file
            logging.info(fullname)
            # copy marks ,add course name            
            wb = openpyxl.load_workbook( fullname)
            sheet = wb.get_active_sheet()
            for row in range(1,sheet.max_row):
                # the first element of every row is file name(include course name)
                rowmark = [file]      # for store marks of one row
                for col in range(1,sheet.max_column):
                    rowmark.append(sheet.cell(row = row, column = col).value)                
                self.coursemark.append(rowmark)
            #logging.info(self.coursemark)
#            input('for debug 1')
            self.allcoursemark.setdefault(classname, [])
            self.allcoursemark[classname].append(self.coursemark)
#            logging.info(self.allcoursemark)            
#        logging.info(self.allcoursemark)   # data is too much to information                   
#        input('for debug 2')
        pass
    
    def copyclass(self, srcdir, dstdir):
        pass
    
    def writeclass(self, dstdir):        
        for classname,mark in self.allcoursemark.items():
            wb = openpyxl.Workbook()            
            sheet = wb.get_active_sheet()
            # Write marks 
            row = 0
            logging.info(len(mark))
            logging.info(len(mark[0]))
            logging.info(len(mark[0][0]))
            for k in range(len(mark)):     # k is begin from 0 to max
                for r in range(len(mark[k])):
                    row += 1
                    for col in range(len(mark[k][r])):
                        sheet.cell(row = row, column = col + 1).value = mark[k][r][col]
                        logging.info(str(row) + ' ' + str(r) + ' ' + str(col) + ': ' + str(mark[k][r][col]))
            newfullname = dstdir + '\\cj-total-' + classname + '.xlsx'
            wb.save(newfullname)
            print('Marks of %s have been written.' % (classname))
#            input('any key') 
            
        pass
    
    def writestudent(self, srcdir, dstdir):
        pass
    def copyall(self, srcdir, dstdir):
        pass
    def writeall(self, srcdir, dstdir):
        pass

    
    # copy all marks in one sheet. add the course name .    
    def allmark(self, srcdir, dstdir):
        self.studentmarks = []        
        for t in self.classlist:
            logging.info(t);logging.info(t[0]);logging.info(t[1])
            classN = t[0]            
            file = t[1]            
            fullname = srcdir + '\\' + file
            logging.info(fullname)
            # copy marks ,add course name            
            wb = openpyxl.load_workbook( fullname)
            sheet = wb.get_active_sheet()
            for row in range(1,sheet.max_row):
                rowmark = []
                rowmark.append(file)
                for col in range(1,sheet.max_column):
                    rowmark.append(sheet.cell(row = row, column = col).value)
                self.studentmarks.append(rowmark)
                logging.info(rowmark)
#                input('for debug')
            logging.info(self.studentmarks)
        # Write marks    
        wb = openpyxl.Workbook()
        sheet = wb.get_active_sheet()
        for row in range(1,len(self.studentmarks)):
            for col in range(1,len(self.studentmarks[0])):
                sheet.cell(row = row, column = col + 1).value = self.studentmarks[row - 1][col - 1]        
        newfullname = dstdir + '\\cj-total' + '.xlsx'
        wb.save(newfullname)

    # copy one class' marks in one sheet. add the course name .    
    def classmark(self, srcdir, dstdir):

        pass


    # collectmarks for students one by one .
    def tidymark(self):

        studmark = [
            ['课程',],
            ['学号', ],
            ['姓名', ],
            ['课堂平时成绩',],
            ['课堂期末成绩',],
            ['课堂总成绩',],
            ['实践成绩',],
            ['实验成绩',],
            ['总成绩',],
            ]
        COLUMNS_MAP = {
                '课堂平时成绩': {'column': 'J', 'index': 3},
                '课堂期末成绩': {'column': 'M', 'index': 4},
                '课堂总成绩': {'column': 'O', 'index': 5},
                '实践成绩': {'column': 'Q', 'index': 6},
                '实验成绩': {'column': 'R', 'index': 7},
                '总成绩': {'column': 'S', 'index': 8},
                }
        
        for test in range(len(self.studentmarks[22])):
            print(self.studentmarks[22][test])     #  column 2 (0,1,2) is student's number.
        input('for debug')
        
        wb = openpyxl.Workbook()
        sheet = wb.get_active_sheet()        

        pass


def main():
    SRCDIR = 'd:\\_PythonWorks\\excelOperate\\cj-2016201701'
    DSTDIR = 'd:\\_PythonWorks\\excelOperate\\cjcl-2016201701'
    RELIST = '-(\d{7}z?)\.' # CLASSREG = re.compile(r'\d{7}[zZ]?')

    collmark = collectmarks()
    CLASSLIST = collmark.classfind(SRCDIR,RELIST)
    
    collmark.copymark(SRCDIR)    

#    collmark.copyclass(SRCDIR,DSTDIR)
    collmark.writeclass(DSTDIR)

#    input('for debug')
    
##    collmark.writestudent(SRCDIR,DSTDIR)
##
##    collmark.copyall(SRCDIR,DSTDIR)
##    collmark.writeall(SRCDIR,DSTDIR)
#     collmark.allmark(self, srcdir, dstdir)   
#    collmark.tidymark()

if __name__ == '__main__':
    logging.critical('--------Start of program---------')
    main()
    logging.critical('--------End---------')
    
