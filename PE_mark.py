#! python 3
# _*_ coding: utf_8  _*_

'''

'''

import openpyxl

import logging

# logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.basicConfig( level = logging.ERROR, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.critical('--------Start of program---------')

MAXROW = 106

China_str = set(['100米','跳远','铅球','８００米','男子成绩','女子成绩'])
PE_items = set(['100米','跳远','铅球','８００米'])
item_head = {}
mark_man = {}
mark_woman = {}

wb_val = {}


wb = openpyxl.load_workbook('高考体育成绩对照表.xlsx')
sheet = wb.get_active_sheet()
for col in range(1,sheet.max_column+1):
    key1 = col
    col_dict = {}
    for row in range(1,sheet.max_row+1):
        key2 = row
        val = sheet.cell( row=row, column=col).value
        col_dict[ key2] = val
    wb_val[ key1] = col_dict
print( key1, key2, val)

for col, row_val in wb_val.items():
    for row, val in row_val.items():
        if val in China_str:
            if val in PE_items:
                item_head[val] = ( col, row)
print(item_head)

for val, ( col, row) in item_head.items():
    key1 = val
    item_man = {}
    item_woman = {}
    for r in range( row+2, MAXROW):
        key_m = wb_val[col][r]
        val_m = wb_val[col+1][r]
        item_man[key_m] = val_m
        key_wo = wb_val[col+2][r]
        val_wo = wb_val[col+3][r]
        item_woman[key_wo] = val_wo
    mark_man[key1] = item_man
    mark_woman[key1] = item_woman
##    print(mark_man)
##    print(mark_woman)
##    input('any key, just for debug')

logging.critical('-------End--------')






