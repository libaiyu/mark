#! python 3
# _*_ coding: utf_8  _*_

'''
During the course, note the performance for the students.
select the "课程"，select "加减分的项",input"学号","分值"
it will write the "分值"
Record can not be written. prompt is not good enough.   2017-2-11-10:50
Student's number must be 2 digits. When it is smaller than 10, It should be 0X.  2017-2-11-17:23
tidy the value. add display the 总分。display the rank.         2017-2-20
debug the course regular.      2017-3-4.
fix the IndexError.     by evan              2017-3-11
add save error capture. add some Tags.       2017-3-12
'''

import openpyxl
import os
import re

from getdir import *  # 2017-3-12

import logging

# logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.basicConfig( level = logging.ERROR, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.critical('--------Start of program---------')

# Chinese words '学生名单' in filename
Chinese_reg = re.compile(r'学生名单')
# class
class_reg = re.compile(r'\d{7}')
# course
course_reg = re.compile(r'-([a-z]{3,11})-')     # 2017-3-4 debug.

performance_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['提出问题',],
    ['回答问题',],
    ['课堂作业',],
    ['作业1',],['作业2',],['作业3',],['作业4',],['作业5',],['作业6',],['作业7',],
    ]
lab_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],['操作5',],['操作6',],['操作7',],['操作8',],
    ['数据1',],['数据2',],['数据3',],['数据4',],['数据5',],['数据6',],['数据7',],['数据8',],   
    ['报告1',],['报告2',],['报告3',],['报告4',],
    ]
design_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['设计1',],['设计2',],['设计3',],['设计4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],  
    ['报告1',],['报告2',],
    ]
practice_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],   
    ['报告1',],['报告2',],    
    ]

tagdict = {'performance':performance_tag,
           'lab':lab_tag,
           'design':design_tag,
           'practice':practice_tag}

# Get the directory name.
DIRNAME = getdir()
# Get the filename list.
FILELIST = os.listdir( DIRNAME)
# Get the select list include coursetype.
filelist = filesele( FILELIST, course_reg)
# sort the filelist. so the index of the file is nochange.
filelist.sort()
# Prompt the number and filename to select.
k = 0
for line in filelist:
    print(k, line)
    k += 1
# File full name list.
fulllist = getfull( DIRNAME, filelist)
st = 'select course或q:'
cour_num = getdigits( st, 0, k)  # input a digit, it is smaller than k.
if cour_num is 'q':
    pass
else:
    coursenum = int( cour_num)
    logging.critical( filelist[coursenum])
    pfrank( fulllist[coursenum], 3)    # print rank.
    coursetype = course_reg.search(filelist[coursenum]).group(1)
    need_check = tagdict[coursetype]
    # need_check = ['提出问题', '课堂作业', '作业1']
    need_check = [ ['作业1',],]
    for each in need_check:
        item_mark( fulllist[coursenum], each[0], 3)  # 分数为0的同学.
    print(coursetype)
    k = 0
    for val in tagdict[coursetype]:
        print(str(k) + ' ' + str(val) + ' ')
        k += 1

    st = 'select item或q:'
    itnum = getdigits( st, 0, k)
    if itnum is 'q':
        pass
    else:
        itemnum = 6 + int( itnum)

        wb = openpyxl.load_workbook(fulllist[coursenum])

        ##try:
        ###    import pdb;pdb.set_trace()
        ##    wb = openpyxl.load_workbook(fulllist[coursenum])
        ##except IndexError:
        ##    input('Please open the workbook, save it and close it.')
        ##    wb = openpyxl.load_workbook(fulllist[coursenum])

        sheet = wb.get_active_sheet()

        finish = itnum
        while finish.isdigit():   # when finish is digit, do loop.
            st = "\n 输入学号或q: 205, 401:"
            studnum = getdigits( st, 100, 900)
            if not studnum.isdigit():
                finish = studnum   # when studnum is not digit, then finish is not digit, means end.
                break
            for row in range(3,sheet.max_row + 1):
                logging.debug(str(sheet['b'+str(row)].value)[-3:])           #  学号在B列
                if str(sheet['b'+str(row)].value)[-3:] == studnum:                   #  学号在B列
                    # Write
                    logging.critical(sheet.cell(row = row,column = itemnum).value)   # 写之前，cell的值
                    mark = input('\n please input the mark: ')
                    sheet.cell(row = row,column = itemnum).value += int(mark)        # 加上要加减的分数
                    sheet.cell(row = row,column = 4).value += int(mark)              # 总分也加上该分数
                    logging.critical(sheet.cell(row = row,column = itemnum).value)   # 写之后，cell的值
                    print(studnum,' 总分是：', sheet.cell(row = row,column = 4).value)
                    break
        while True:
            try:    
                wb.save(fulllist[coursenum])
            except PermissionError:
                input('Please close the workbook.')
            else:
                break
        pfrank( fulllist[coursenum], 8)    # print rank.

logging.critical('-------End--------')






