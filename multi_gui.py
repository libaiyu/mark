
#! python3
# _*_ coding: utf_8  _*_
'''list the ahead 8 students' mark.        2017-3-5
try to list the courses. then we can choose the course by click.  2017-3-6
try to arrange the widgets.                2017-3-7
try to add a Listbox.  it can work. but not strong. minus can not work. 2017-3-15
Minus mark can be written now.             2017-3-16  morning
It can be run arbitrary.                   2017-3-16  22:10

多页面                2017-3-26
增加查询作业上交情况前的输入提示    2017-3-28
''' 

import os
import re
from tkinter import *
import tkinter as tk
from tkinter import ttk

import openpyxl
import matplotlib
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2TkAgg
from matplotlib.figure import Figure

from getdir import getdir, filesele, getfull, getbackup

# course type
COURSE_REG = re.compile(r'-([a-z]{3,11})-')     # 2017-3-4 debug.

global CONTENTS            # global variable

PERFORMANCE_TAG = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['提出问题',],
    ['回答问题',],
    ['课堂作业',],
    ['作业1',],['作业2',],['作业3',],['作业4',],['作业5',],['作业6',],['作业7',],['是否已交课堂作业',],
    ]
LAB_TAG = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],['操作5',],['操作6',],['操作7',],['操作8',],
    ['数据1',],['数据2',],['数据3',],['数据4',],['数据5',],['数据6',],['数据7',],['数据8',],   
    ['报告1',],['报告2',],['报告3',],['报告4',],
    ]
DESIGN_TAG = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['设计1',],['设计2',],['设计3',],['设计4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],  
    ['报告1',],['报告2',],
    ]
PRACTICE_TAG = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],   
    ['报告1',],['报告2',],    
    ]
TAGDICT = {'performance':PERFORMANCE_TAG,
           'lab':LAB_TAG,
           'design':DESIGN_TAG,
           'practice':PRACTICE_TAG}

##import tkinter as tk
##from tkinter import ttk

##import matplotlib
matplotlib.use("tkagg")

##from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2TkAgg
##from matplotlib.figure import Figure


LARGE_FONT= ("Verdana", 12)


class Application(tk.Tk):
    '''
    平时成绩管理-多页面版
        界面与逻辑分离
    '''
    def __init__(self):
        
        super().__init__()

        self.iconbitmap(default="whoami.ico")
        self.wm_title("平时成绩管理-多页面版")
        
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand = True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}
        for F in (StartPage, PageOne, PageTwo, PageThree):
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")  # 四个页面的位置都是 grid(row=0, column=0), 位置重叠！！

        self.show_frame(StartPage)
        
    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise() # 切换，提升当前 tk.Frame z轴顺序（使可见）！！此语句是本程序的点睛之处

        
class StartPage(tk.Frame):
    '''选课程'''
    def __init__(self, parent, root):
        super().__init__(parent)
        label = tk.Label(self, text="选课程", font=LARGE_FONT)
        label.pack()
        
        button1 = ttk.Button(self, text="去到课后", command=lambda: root.show_frame(PageTwo)).pack()
        button2 = ttk.Button(self, text="去到成绩录入", command=lambda: root.show_frame(PageOne)).pack()
##        button3 = ttk.Button(self, text="去到其他", command=lambda: root.show_frame(PageThree)).pack()

        self.create_widgets()

    def create_widgets(self):
        'Create all kinds of widgets.'

        '''界面'''
        self.mainframe = ttk.Frame(self, padding="9 9 12 12") # 注意ttk.Frame()的第一个参数为self，因为这个类继承自tk.Tk类
        self.mainframe.pack()
        self.mainframe.columnconfigure(0, weight=1)
        self.mainframe.rowconfigure(0, weight=1)

        # Create"选择课程" button .
        Button( self.mainframe, text = "选择课程", command=self.sele_course).pack()
        
        # label.
        Label( self.mainframe, text = '所选课程').pack()
        
        global CONTENTS             # global variable
        CONTENTS = StringVar()
        CONTENTS.set('用于显示所选择的课程.每页都相同。')        # set it to some value
        
        # Entry.
        self.course_ent = Entry( self.mainframe, width=85, textvariable = CONTENTS)
        self.course_ent.pack()

        # "旷课者查询" button .
        Button(self.mainframe, text = "旷课者查询", command = self.find_absent).pack()

        # Create listbox.
        self.listbox =  Listbox(self.mainframe, width=85, height=15)
        self.listbox.pack()

        # "前8名:" button .
        Button(self.mainframe, text = "前8名:", command = self.ahead).pack()

        # Create"列出课程" button .
        Button( self.mainframe, text = "列出课程", command=self.listfile).pack()
        
        # '输入分数' label.
        Label( self.mainframe, text = '输入学号后3位。').pack()
        
        self.cont = StringVar()
        self.cont.set('请输入学号后3位：')   # set it to some value
        # Entry 2.
        self.look_ent = Entry( self.mainframe, width=85, textvariable = self.cont)
        self.look_ent.pack()
        self.look_ent["textvariable"] = self.cont    # tell the entry widget to watch this variable
        self.look_ent.bind('<Key-Return>', self.lookmark)  # when the user hits return

        # "旷课者查询" button .
        Button(self.mainframe, text = "旷课者查询", command = self.find_absent).pack()
        
        # Create"清空列表框" button .
        Button( self.mainframe, text = "清空列表框", command=self.clrlistbox).pack()

    def listfile(self):     #  "列出课程"
        
        global fulllist
##        self.listbox.delete( 0, END)
        self.listbox.insert( 0, '\n')
        for coursen in fulllist:
            self.listbox.insert( 0, coursen)
        self.listbox.insert( 0, '\n')

    def sele_course(self):    # "选择课程"
        
        global fulllist, NUM
        if NUM <= 0:
            NUM = len( fulllist)
        if NUM > 0:
            # every click, NUM decrease 1. to select the next course.
            NUM -= 1       #  decrease must be place here.
            CONTENTS.set( fulllist[NUM])
            self.listbox.insert( 0, fulllist[NUM])
            
    def find_absent( self):  # 找出旷课者

        global fulllist, NUM
        # Open the book.
        col = 19        # column 19 is "是否上交课堂作业"
        wb = openpyxl.load_workbook( fulllist[NUM])
        sheet = wb.get_active_sheet()
        ab_stu = []
        for row in range(3,sheet.max_row + 1):
            if not sheet.cell( row=row,column=col).value:
                ab_stu.append( sheet[ 'b'+str( row)].value)
        wb.save( fulllist[NUM])
        
        # self.listbox.delete( 0, END)
        if len( ab_stu)<5:   # len( ab_stu) = 0 时，会插入一个空列，感觉比什么都没有更踏实。
            self.listbox.insert( 0, ab_stu)
        else:    #   len( ab_stu)>=5
            for k in range( len( ab_stu)//5):
                self.listbox.insert( 0, ab_stu[ 5*k:5*k+5])
            if len( ab_stu)%5:
                self.listbox.insert( 0, ab_stu[ 5*k+5:])
        pass
        self.listbox.insert( 0, fulllist[NUM]+'  旷课者：\n')

    def ahead(self):   # display the ahead 8 marks.
        
        global fulllist, NUM
        # Read the marks.
        wb = openpyxl.load_workbook( fulllist[NUM])
        sheet = wb.get_active_sheet()
        marks = []
        for row in range(3,sheet.max_row + 1):
            marks.append((sheet.cell(row = row,column = 4).value, sheet['b'+str(row)].value, sheet['c'+str(row)].value))
        wb.save( fulllist[NUM])
        
        # self.list_box.delete(1.0, END)
##        self.listbox.insert( 0, '前8名为：\n')
        # rank the marks.
        marks.sort(reverse=True)
        for n in range( len( marks[:8])-1, -1, -1):
            # insert the ahead marks to text.
           self.listbox.insert( 0, str(marks[n])) #  self.listbox.insert( 0, str(marks[n])+'\n')

        self.listbox.insert( 0, fulllist[NUM]+'  前8名为：\n')
        pass

    def lookmark( self, event):
        ''' Look up the marks of given students' numbers.  '''
        
        global fulllist, NUM

        # Read the Entry. Then backup.
        st = self.look_ent.get()   #  multi student number split by ",".
        print( st)
        backup = 'y'  #  input('Is this need backup?')
        if backup.lower() == 'y':
            import datetime
            global BACKUPFILE
            t = datetime.datetime.now()
            memory_file = open( BACKUPFILE,'a')
            memory_file.write( '\n'+str( t.year)+'-'+str( t.month)+'-'+str( t.day)+','+str( t.hour)+':'+str( t.minute)+':'+str( t.second)+'\n')
            memory_file.write( fulllist[NUM]+'\n')
            memory_file.write( st+'\n')
            memory_file.close()

        # Split each student number. Form the students' number list.
        mm = st.split(',')
        # get the studnum
        studnum = []
        for e in range(len(mm)):
            if not mm[e].isdigit():    # each student number should be digit.
                break                  # till it is not digit, break.

            if len(mm[e]) == 3:          # student number.
##                print(mm[e],type(mm[e]))
                studnum.append( mm[e])
                class_temp = mm[e][0]
            elif len(mm[e]) == 2:        # student number.
                mmm = class_temp + mm[e]
                studnum.append( mmm)
            else:
                print('数字位数不对。')
        pass
        print( fulllist[NUM])
        print( studnum)

        self.listbox.insert( 0, fulllist[NUM])
        # open the excel file.
        wb = openpyxl.load_workbook(fulllist[NUM])
        sheet = wb.get_active_sheet()
        # Read the mark
        result = []
        # Read the title
        row_title = []
        for col in range( 2, sheet.max_column+1):
            row_title.append( sheet.cell( row=2, column=col).value)
        result.append( row_title)

        # Read the marks.
        for e in range(len(studnum)):
            for row in range(3,sheet.max_row + 1):
                if str( sheet[ 'b'+str( row)].value)[-3:] == studnum[e]:           #  学号在B列
                    # Look up the mark.
                    row_m = []
                    for col in range( 2, sheet.max_column+1):
                        row_m.append( sheet.cell( row=row, column=col).value)
                    result.append( row_m)
                    break
        print( result)
        for each in result:
            self.listbox.insert( 0, each)

        pass
    
    def clrlistbox(self):
        self.listbox.delete(0, END)
        pass
    

class PageOne(tk.Frame):
    '''成绩录入'''
    def __init__(self, parent, root):
        super().__init__(parent)
        label = tk.Label(self, text="这是成绩录入", font=LARGE_FONT)
        label.pack()

        button1 = ttk.Button(self, text="回到选课程", command=lambda: root.show_frame(StartPage)).pack()
        button2 = ttk.Button(self, text="去到课后", command=lambda: root.show_frame(PageTwo)).pack()

        self.create_widgets()

    def create_widgets(self):
        'Create all kinds of widgets.'

        '''界面'''
        self.mainframe = ttk.Frame(self, padding="9 9 12 12") # 注意ttk.Frame()的第一个参数为self，因为这个类继承自tk.Tk类
        self.mainframe.pack()
        self.mainframe.columnconfigure(0, weight=1)
        self.mainframe.rowconfigure(0, weight=1)

        # '所选课程' label .
        Label( self.mainframe, text = '所选课程').pack()
        
        # Entry 1.
        self.course_ = Entry( self.mainframe, width=85, textvariable = CONTENTS)
        self.course_.pack()

        # "排名:" button .
        Button(self.mainframe, text = "排名:", command = self.rank).pack()
        
        # "显示" listbox.
        self.list_box = Listbox(self.mainframe, height=12, width=88)
        self.list_box.pack()

        # Create"清空列表框" button .
        Button( self.mainframe, text = "清空排名－列表框", command=self.clrlistbox).pack()
        
        # "清除上次课堂作业上交记录" button .
        Button(self.mainframe, text = "清除上次课堂作业上交记录", command = self.clr_absent).pack()

        # "加减分项目" button .
        Button(self.mainframe, text = "加减分项目", command = self.list_item).pack()

        # "加减分项目" Text.
        self.itemtext = Text(self.mainframe, width=85, height=3, wrap='word')
        self.itemtext.pack()
        
        # '输入分数' label.
        Label( self.mainframe, text = '输入分数').pack()
        
        self.cont = StringVar()
        self.cont.set('请输入项目代号2位学号3位及分数-2分记为12：0511102')   # set it to some value
        # Entry 2.
        self.mark_ent = Entry( self.mainframe, width=85, textvariable = self.cont)
        self.mark_ent.pack()
        self.mark_ent["textvariable"] = self.cont    # tell the entry widget to watch this variable
        self.mark_ent.bind('<Key-Return>', self.markin)  # when the user hits return

        # "旷课者查询" button .
        Button(self.mainframe, text = "旷课者查询", command = self.find_absent).pack()

        # Create"清空文本框" button .
        Button( self.mainframe, text = "清空项目－文本框", command=self.clrtext).pack()

    def list_item(self):
        
        global fulllist, NUM
        
        # list the item according to the coursetpye.
##        self.itemtext.delete( 0, END) # self.list_box.delete(0.0, END)
        coursetype = COURSE_REG.search( fulllist[NUM]).group(1)
        k = 0
        for val in TAGDICT[coursetype]:
            self.itemtext.insert( END, str(k)+','+str(val)+'\t\t')
            k += 1
        self.cont.set('0512702')
        pass
    
    def markin( self, event):    #    event.??    ok.   17-3-15.
        # mark update.
        
        global fulllist, NUM

        # get the multi marks from Entry 2.
        st = self.mark_ent.get()   #  multi marks split by ",".
        print( st)
        backup = 'y'  #  input('Is this need backup?')
        if backup.lower() == 'y':
            import datetime
            global BACKUPFILE
            t = datetime.datetime.now()
            memory_file = open( BACKUPFILE,'a')
            memory_file.write( str( t.year)+'-'+str( t.month)+'-'+str( t.day)+','+str( t.hour)+':'+str( t.minute)+':'+str( t.second)+'\n')
            memory_file.write( fulllist[NUM]+'\n')
            memory_file.write( st+'\n')
            memory_file.close()

        # split each mark.
        mm = st.split(',')
        # get the itemnum, studnum, mark list.
        itemnum = []
        studnum = []
        mark = []
        for e in range(len(mm)):
            if not mm[e].isdigit():    # each mark should be digit. if so, it will write in.
                break                  # till mark is not digit, break.

            if len(mm[e]) == 7:          # item, student number(3 digits), mark.
##                print(mm[e],type(mm[e]))
                ittemp = 6 + int( mm[e][:2])
                itemnum.append( ittemp)
                studnum.append( mm[e][2:5])
                marktemp = mm[e][5:]
                mark.append( marktemp)
##            if len(mm[e]) == 6:          # item, student number(2 digits), mark.
##                ittemp = 6 + int( mm[e][:2])
##                itemnum.append( ittemp)
##                studnum.append( mm[e][2:4])
##                marktemp = mm[e][4:]
##                mark.append( marktemp)
            elif len(mm[e]) == 5:        # item no change, student number change, mark change.
                itemnum.append( ittemp)   # item no change.
                studnum.append( mm[e][:3])
                marktemp =  mm[e][3:]
                mark.append( marktemp)
            elif len(mm[e]) == 3:        # item no change, student number change, mark no change.
                itemnum.append( ittemp)
                studnum.append( mm[e])
                mark.append( marktemp)   #  mark no change.
            else:
                print('数字位数不对。')
        pass
        print( fulllist[NUM])
        print( itemnum)
        print( studnum)
        print( mark)

        self.list_box.insert( 0, fulllist[NUM])
        # open the excel file.
        wb = openpyxl.load_workbook(fulllist[NUM])
        sheet = wb.get_active_sheet()

        # write the marks.
        for e in range(len(studnum)):
            for row in range(3,sheet.max_row + 1):
                if str( sheet[ 'b'+str( row)].value)[-3:] == studnum[e]:           #  学号在B列
                    # Write
                    if itemnum[e] == 11:      # 课堂作业已做，说明来上课了。
                        if sheet.cell(row=row,column=19).value:
                            re_in = input('已经记录了课堂作业成绩，还要增加记录吗？')
                            if re_in == 'y':
                                pass
                            else:
                                break
                        sheet.cell(row=row,column=19).value += 1       # 上课记录，0表示缺课,1 is nomal, 2 means repeat write.
                            
                    elif sheet.cell(row = row,column = itemnum[e]).value:
                        re_in = input('已经记录了作业成绩，还要记录吗？')
                        if re_in == 'y':
                            pass
                        else:
                            break
                    else:
                        pass
                    dp = str( sheet[ 'b'+str( row)].value) +':'+ '  ' + str( sheet.cell(row = row,column = itemnum[e]).value)
                    dp1 = '  ' + str( sheet.cell(row = row,column = 4).value)+'->'
##                    self.list_box.insert( 0, dp+' ')                       # 显示加之前的分数
                    if mark[e][:1]=='0':
                        sheet.cell(row = row,column = itemnum[e]).value += int(mark[e][1:])        # 加上要加的分数
                        sheet.cell(row = row,column = 4).value += int(mark[e][1:])              # 总分也加上该分数
                    elif mark[e][:1]=='1':
                        sheet.cell(row = row,column = itemnum[e]).value -= int(mark[e][1:])    # 减去要减的分数
                        sheet.cell(row = row,column = 4).value -= int(mark[e][1:])          # 总分也减去该分数
                    dp += '->'+ str( sheet.cell( row = row,column = itemnum[e]).value)
                    dp1 += str( sheet.cell(row = row,column = 4).value)
                    dp += dp1
                    self.list_box.insert( 0, dp+'  ')                     # 显示加之前加之后的分数   0  END
                    break
        pass

        # save the excel file.
        while True:
            try:    
                wb.save(fulllist[NUM])
            except PermissionError:
                input('Please close the workbook.')
            else:
                break
        pass
    
    def rank(self):
        
        global fulllist, NUM
        # Read the marks.
        wb = openpyxl.load_workbook( fulllist[NUM])
        sheet = wb.get_active_sheet()
        marks = []
        for row in range(3,sheet.max_row + 1):
            marks.append((sheet.cell(row = row,column = 4).value, sheet['b'+str(row)].value, sheet['c'+str(row)].value))
        wb.save( fulllist[NUM])
        
##        self.list_box.delete(1.0, END)
        # rank the marks.
        marks.sort()
        for k in marks[:]:
            # insert the ahead marks to text.
            self.list_box.insert(0, str(k)+'\n')
        self.list_box.insert(0, fulllist[NUM]+'  排名为：\n')
        pass

    def find_absent( self):

        global fulllist, NUM
        # Open the book.
        col=19        # column 12 is "是否上交课堂作业"
        wb = openpyxl.load_workbook( fulllist[NUM])
        sheet = wb.get_active_sheet()
        ab_stu = []
        for row in range(3,sheet.max_row + 1):
            if not sheet.cell( row=row,column=col).value:
                ab_stu.append( sheet[ 'b'+str( row)].value)
        wb.save( fulllist[NUM])
        
##        self.list_box.delete( 0, END)
##        self.list_box.insert( 0, '\n')
        if len( ab_stu)<5:   # len( ab_stu) = 0 时，会插入一个空列，感觉比什么都没有更踏实。
            self.list_box.insert( 0, ab_stu)
        else:    #   len( ab_stu)>=5
            for k in range( len( ab_stu)//5):
                self.list_box.insert( 0, ab_stu[ 5*k:5*k+5])
            if len( ab_stu)%5:
                self.list_box.insert( 0, ab_stu[ 5*k+5:])
        pass
        self.list_box.insert( 0, fulllist[NUM]+'  旷课者：\n')

    def clr_absent(self):

        global fulllist, NUM
        # Open the book.
        wb = openpyxl.load_workbook( fulllist[NUM])
        sheet = wb.get_active_sheet()
        for row in range(3,sheet.max_row + 1):
            sheet.cell(row=row,column=19).value = 0
        wb.save( fulllist[NUM])
        pass

    def clrlistbox(self):
        self.list_box.delete( 0, END)
        pass
    
    def clrtext(self):
        self.itemtext.delete( 0.0, END)
        pass

    
class PageTwo(tk.Frame):
    '''课后'''
    def __init__(self, parent, root):
        super().__init__(parent)
        label = tk.Label(self, text="这是课后", font=LARGE_FONT)
        label.pack()

        button1 = ttk.Button(self, text="回到选课程", command=lambda: root.show_frame(StartPage)).pack()
        button2 = ttk.Button(self, text="回到成绩录入", command=lambda: root.show_frame(PageOne)).pack()

        self.create_widgets()

    def create_widgets(self):
        'Create all kinds of widgets.'

        '''界面'''
        self.mainframe = ttk.Frame(self, padding="9 9 12 12") # 注意ttk.Frame()的第一个参数为self，因为这个类继承自tk.Tk类
        self.mainframe.pack()
        self.mainframe.columnconfigure(0, weight=1)
        self.mainframe.rowconfigure(0, weight=1)

        # label.
        Label( self.mainframe, text = '所选课程').pack()

        # Entry.
        self.course_ent = Entry( self.mainframe, width=85, textvariable = CONTENTS)
        self.course_ent.pack()

        # "提示输入第几次作业" label .
        Label(self.mainframe, text = "请输入要查询第几次作业").pack()

        self.whichhw = StringVar()
        self.whichhw.set('3')   # 2 means the second homework.
        # Entry 3.
        self.homew_ent = Entry( self.mainframe, width=3, textvariable = self.whichhw)
        self.homew_ent.pack()
        
        # "作业已做者查询" button .
        Button(self.mainframe, text = "作业已做者查询", command = self.homeworkck).pack()
        
        # "作业未做者查询" button .
        Button(self.mainframe, text = "作业未做者查询", command = self.nohomework).pack()
        
        # '查询结果' listbox.
        self.list_box = Listbox(self.mainframe, width=85, height=15)
        self.list_box.pack()

        # "排名:" button .
        Button(self.mainframe, text = "排名:", command = self.rank).pack()
        
##        # text.    # text display is not good as listbox
##        self.list_box = Text(self.mainframe, height=15, width=85, wrap='word')
##        self.list_box.pack()
        
        # Create"清空列表框" button .
        Button( self.mainframe, text = "清空列表框", command=self.clrlistbox).pack()

    def homeworkck( self):
        
        global fulllist, NUM
        
        # which homework.
        which = self.homew_ent.get()
        
        col= int(which) + 11        # column 12 is "作业1"
        # Open the book.
        wb = openpyxl.load_workbook( fulllist[NUM])
        sheet = wb.get_active_sheet()
        homework = []
        for row in range(3,sheet.max_row + 1):
            if sheet.cell( row=row,column=col).value:
                homework.append( sheet[ 'b'+str( row)].value)
##        try:
        wb.save( fulllist[NUM])
##        except KeyError:
##            # print('please input a number for the homework you want to check.')
##            self.course.set('please input a number for the homework you want to check.')
        
        # display the student number who have not submit the homework.
        
##        self.list_box.insert( 0, '\n')
        if len( homework) < 5:
            self.list_box.insert( 0, homework)
        else:
            for k in range( len( homework)//5):
                self.list_box.insert( 0, homework[ 5*k:5*k+5])
            if len( homework)%5:
                self.list_box.insert( 0, homework[ 5*k+5:])
        self.list_box.insert( 0, fulllist[NUM]+'  第'+ which+'次作业 已做者：\n')        
        pass
        
    def nohomework( self):

        global fulllist, NUM

##        self.course.delete( 0, END)
        # input('please input a number for the homework you want to check.')  # how to prompt .
        which = self.homew_ent.get()
##        self.cont.set["textvariable"] = '第几次作业？which ='
##        which = int( input( '第几次作业？'))
##        self.course.insert( 0, which)
##        self.list_box.insert( 0, which)
        col= int(which) + 11        # column 12 is "作业1"
        # Open the book.
        wb = openpyxl.load_workbook( fulllist[NUM])
        sheet = wb.get_active_sheet()
        nohome = []
        for row in range(3,sheet.max_row + 1):
            if not sheet.cell( row=row,column=col).value:
                nohome.append( sheet[ 'b'+str( row)].value)
##        try:
        wb.save( fulllist[NUM])
##        except KeyError:
##            # print('please input a number for the homework you want to check.')
##            self.course.set('please input a number for the homework you want to check.')
        
        # display the student number who have not submit the homework. 

##        if len( nohome) < 5:
##            self.course.insert( 0, nohome)
##        else:
##            for k in range( len( nohome)//5):
##                self.course.insert( 0, nohome[ 5*k:5*k+5])
##            if len( nohome)%5:
##                self.course.insert( 0, nohome[ 5*k+5:])

##        self.list_box.insert( 0, '\n')
        if len( nohome) < 5:
            self.list_box.insert( 0, nohome)
        else:
            for k in range( len( nohome)//5):
                self.list_box.insert( 0, nohome[ 5*k:5*k+5])
            if len( nohome)%5:
                self.list_box.insert( 0, nohome[ 5*k+5:])
        self.list_box.insert( 0, fulllist[NUM]+'  第'+ which+'次作业 未做者：\n')        
        pass
    
    def clrlistbox(self):
        self.list_box.delete( 0, END)
        pass

    def rank(self):
        
        global fulllist, NUM
        # Read the marks.
        wb = openpyxl.load_workbook( fulllist[NUM])
        sheet = wb.get_active_sheet()
        marks = []
        for row in range(3,sheet.max_row + 1):
            marks.append((sheet.cell(row = row,column = 4).value, sheet['b'+str(row)].value, sheet['c'+str(row)].value))
        wb.save( fulllist[NUM])
        
##        self.list_box.delete(1.0, END)
        # rank the marks.
        marks.sort()
        for k in marks[:]:
            # insert the ahead marks to text.
            self.list_box.insert(0, str(k)+'\n')
        self.list_box.insert(0, fulllist[NUM]+'  排名为：\n')
        pass


class PageThree(tk.Frame):
    '''其他'''
    def __init__(self, parent, root):
        super().__init__(parent)
        tk.Label(self, text="这是其他", font=LARGE_FONT).pack()

        button1 = ttk.Button(self, text="回到选课程", command=lambda: root.show_frame(StartPage)).pack()


def getfile():

    global fulllist, NUM, BACKUPFILE, f
    
    # Get the directory name.
    DIRNAME = getdir()
    BACKUPFILE = getbackup()
    # Get the filename list.
    FILELIST = os.listdir( DIRNAME)
    # Get the filename list include coursetype.
    filelist = filesele( FILELIST, COURSE_REG)
    # sort the filelist. so the index of the file is nochange.
    filelist.sort()       
    # File full name list.
    fulllist = getfull( DIRNAME, filelist)
    NUM = len( fulllist)

##def clrlistbox( box):
##    box.delete(0, END)
##    pass
    

if __name__ == '__main__':  #  __main__ is not correct.

    # First of all, we get the directory and the excel workbook.
    getfile()
    
    # 实例化Application
    APP = Application()
    
    # 主消息循环:
    APP.mainloop()
    
    print('End')

