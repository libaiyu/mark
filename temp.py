#! python3
# _*_ coding: utf_8  _*_
# copy the remark from a temp workbook, then write to the remark workbook.

import openpyxl
import os
import re

num = []
mark = []
wb = openpyxl.load_workbook('d:\\_PythonWorks\\excelOperate\\bkcj-201702\\BKmark-1520603-4.xlsx')  # BKmark-1520603-4.
sheet = wb.get_active_sheet()
for row in range(2,sheet.max_row+1):
    num.append(str(sheet.cell(row=row, column=2).value))
    mark.append(str(sheet.cell(row=row, column=3).value))
wb.save('d:\\_PythonWorks\\excelOperate\\bkcj-201702\\BKmark-1520603-4.xlsx') # 
input('anything')

	

wb = openpyxl.load_workbook('d:\\_PythonWorks\\excelOperate\\bkcj-201702\\BK_1520603-4.xlsx')
sheet = wb.get_active_sheet()
for n in range(len(num)):
    for row in range(2,sheet.max_row+1):
        if str(sheet['c'+str(row)].value)[-3:] == num[n]:
            sheet['f'+str(row)].value = mark[n]
            
            

wb.save('d:\\_PythonWorks\\excelOperate\\bkcj-201702\\BK_1520603-4.xlsx')

