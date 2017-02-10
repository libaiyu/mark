# _*_ coding: utf_8  _*_
 
import logging
import os
import openpyxl
import pprint
import re

# logging.disable(logging.CRITICAL)
# logging.basicConfig( filename='loglearn.txt',level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )                   
logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.critical('--------Start of program---------')

# Write marks
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('example')
# The student's marks need to write in the cell of the excel
# List is reasonable at here.
studentsMarks ={'学号':'201510040101',
                 '课堂平时成绩':85,
                 '课堂期末成绩':76,
                 '课堂总成绩':76,
                 '实践成绩':76,
                 '实验成绩':76,
                 '总成绩':76,    }

# Loop for write
col = 2
for key in studentsMarks.keys():
    logging.info(key)
    sheet.cell(row = 2,column = col).value = key
    col +=1
    # sheet['str(row)+str(2)'] = key
col = 2
for val in studentsMarks.values():
    logging.info(val)
    sheet.cell(row = 3,column = col).value = val
    col +=1

wb.save('example.xlsx')
 
