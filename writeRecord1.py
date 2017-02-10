# _*_ coding: utf_8  _*_
 
import openpyxl
import os
import re
import logging


logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.critical('--------Start of program---------')

# find the file that include '学生名单' in filename
ChineseReg = re.compile(r'学生名单')
# find the class
classReg = re.compile(r'\d{7}')
# find the course
courseReg = re.compile(r'-(\w{3,11})-')

performanceTag = [
    ['学号', ],
    ['姓名', ],
    ['初始分', ],
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['提出问题',],
    ['回答问题',],
    ['作业1',],['作业2',],['作业3',],['作业4',],['作业5',],['作业6',],['作业7',],['作业8',],
    ]
labTag = [
    ['学号', ],
    ['姓名', ],
    ['初始分', ],
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],['操作5',],['操作6',],['操作7',],['操作8',],
    ['数据1',],['数据2',],['数据3',],['数据4',],['数据5',],['数据6',],['数据7',],['数据8',],
    ['报告1',],['报告2',],['报告3',],['报告4',],    
    ]
designTag = [
    ['学号', ],
    ['姓名', ],
    ['初始分', ],
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['设计1',],['设计2',],['设计3',],['设计4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],
    ['报告1',],['报告2',],    
    ]
practiceTag = [
    ['学号', ],
    ['姓名', ],
    ['初始分', ],
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],
    ['报告1',],['报告2',],    
    ]

# 提示输入信息

fileN = []
k = 0
for fileName in os.listdir('d:\\_PythonWorks\\execlOperate\\pscj161702'):
    fileN.append(fileName)    
    print(k,fileName)
    k += 1
courseType = ['1 理论performance','2 实验lab','3 课程设计design','4 认识实习practice']
print(courseType[0])
performanceItem = []
k = 2
for val in performanceTag:
    performanceItem.append(str(k) + ' ' + str(val) + ' ')
    k += 1
print(performanceItem)
print(courseType[1])
labItem = []
k = 2
for val in labTag:
    labItem.append(str(k) + ' ' + str(val) + ' ')
    k += 1
print(labItem)
print(courseType[2])
designItem = []
k = 2
for val in designTag:
    designItem.append(str(k) + ' ' + str(val) + ' ')
    k += 1
print(designItem)
print(courseType[3])
practiceItem = []
k = 2
for val in practiceTag:
    practiceItem.append(str(k) + ' ' + str(val) + ' ')
    k += 1
print(practiceItem)

itemNum = input('please input numbers for select item: ')


wb = openpyxl.load_workbook(fileN[int(itemNum[:2])])
logging.critical(fileN[int(itemNum[:2])])
sheet = wb.get_active_sheet()
for row in range(1,60):
    logging.debug(str(sheet['b'+str(row)].value)[-2:])           #  学号在B列
    if str(sheet['b'+str(row)].value)[-2:] == itemNum[4:]:                      #  学号在B列
        # Write
        logging.debug(sheet.cell(row = row,column = int(itemNum[2:4])).value)
        sheet.cell(row = row,column = int(itemNum[2:4])).value += 1
        logging.debug(sheet.cell(row = row,column = int(itemNum[2:4])).value)
        wb.save(fileN[int(itemNum[:2])])
        print('one cell is written!')
        break

logging.critical('-------End--------')






