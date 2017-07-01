#! python3
# Get the directory name in this module.
import os


def getdir():
    ''' Get the directory path.

If for test, the directory path will read from a test.txt.

If for work, the directory paht will read from a work.txt.
'''
    global f
    f = input('t(test for testing) or w(work for working):')
    dirn = os.getcwd()
    FILE = dirn+'\\directory_' + f + '.txt'
    # FILE = 'd:\\_PythonWorks\\mark\\directory' + f + '.txt'
    try:
        dirname = open(FILE).read()
    except IOError:
        print('No such filename.')
        dirname = input('Please input the directory name:')
        file = open(FILE, 'w')
        file.write(dirname)
        file.close()
        print('Directory name has been written in ', FILE)
        pass
    else:
        print('We have got directory name', dirname, 'in', FILE, '\n')
        pass
    return dirname


def getbackup():
    ''' Get the file name that used to backup the content of entry.
'''
    backupfile = 'backup_entry_'+f+'.txt'
    return backupfile


def filesele(filelist, regex):
    '''select the file name that according to the regular express.
    '''
    fileselect = []
    for file in filelist:
        # slect the files that not fit regex.
        if regex.search(file):
            fileselect.append(file)
    return fileselect


def getfull(dirname, filelist):
    ''' Form the list of full name.
'''
    full_list = []
    for file in filelist:
        fullname = dirname + '\\' + file
        full_list.append(fullname)
    return full_list


def getdigits(s, min_, max_):
    '''Make the input is the digit or 'q'. s is prompt.
    '''
    while True:
        inum = input('\n'+s+'between: '+str(min_)+', '+str(max_)+':')
        while not inum.isdigit():
            if inum == 'q':
                return inum
            else:
                inum = input('\n 输入的必须是数字或q！')
        itnum = int(inum)
        
        if itnum in range(min_, max_):
            return inum
        else:
            print('\n 注意输入数字的范围：')
            pass


def prank(marks, num):
    '''print the students' mark according to given number.
    '''
    marks.sort(reverse=True)
    print('前'+str(num)+'名为：')
    for k in marks[:num]:
        print(k)
    print('后'+str(num)+'名为：')
    for k in marks[-num:]:
        print(k)


def pfrank(file, num):
    '''print the students' mark according to given number.
    '''
    import openpyxl
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_active_sheet()
    marks = []
    for row in range(3, sheet.max_row + 1):
        mark_sum = sheet.cell(row=row, column=4).value
        stud_num = sheet['b'+str(row)].value
        stud_name = sheet['c'+str(row)].value
        marks.append((mark_sum, stud_num, stud_name))
    while True:
        try:
            wb.save(file)
        except PermissionError:
            input('Please close the workbook.')
        else:
            break
    prank(marks, num)


def item_mark(file, st, num=3):
    '''print students' number and name whose item is zero.
    '''
    import openpyxl
    
    no_flag = 0    # no found flag = 0
    
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_active_sheet()
    for colu in range(1, sheet.max_column+1):
        # print(st)
        if sheet.cell(row=2, column=colu).value == st:
            col = colu
            col_flag = 1    # found flag =1.
    if no_flag:
        print(sheet.cell(row=2, column=col).value)
        print('分数为0的同学有：')
        marks = []
        for row in range(3, sheet.max_row + 1):
            stud_num = sheet['b'+str(row)].value
            stud_name = sheet['c'+str(row)].value
            el = (sheet.cell(row=row, column=col).value, stud_num, stud_name)
            marks.append(el)
            if not sheet.cell(row=row, column=col).value:
                print(el)
        prank(marks, num)
    wb.save(file)

    
def getfile():

    # Get the directory name.
    DIRNAME = getdir()
    BACKUPFILE = getbackup()
    # Get the filename list.
    FILELIST = os.listdir(DIRNAME)
    # Get the filename list include coursetype
    
    import re
    course_reg = re.compile(r'-([a-z]{3,11})-')
    filelist = filesele(FILELIST, course_reg)
    # sort the filelist. so the index of the file is nochange.
    filelist.sort()
    # File full name list.
    fulllist = getfull(DIRNAME, filelist)
    
    return fulllist


def getselefile():

    # Get the directory name.
    DIRNAME = getdir()
    BACKUPFILE = getbackup()
    # Get the filename list.
    FILELIST = os.listdir(DIRNAME)
    # Get the filename list include coursetype
    
    import re
    course_reg = re.compile(r'-([a-z]{3,11})-')
    filelist = filesele(FILELIST, course_reg)
    # sort the filelist. so the index of the file is nochange.
    filelist.sort()
    
    return filelist


def test():

    # Get the directory name.
    DIRNAME = getdir()
    print(DIRNAME)
    input('debug')
    BACKUPFILE = getbackup()
    print(BACKUPFILE)
    input('debug')
    # import os
    # Get the filename list.
    FILELIST = os.listdir(DIRNAME)
    k = 0
    for file in FILELIST:
        print(k, file)
        k += 1
    print('\n')
    import re
    course_reg = re.compile(r'-([a-z]{3,11})-')
    filese = filesele(FILELIST, course_reg)
    k = 0
    for file in filese:
        print(k, file)
        k += 1
    print('\n')
    fulllist = getfull(DIRNAME, filese)
    k = 0
    for line in fulllist:
        print(k, line, end='')
        k += 1
    st = 'please input a number '
    getdigits(st, -100, 300)


if __name__ == '__main__':
    getfile()
    input('-----------')
    test()
