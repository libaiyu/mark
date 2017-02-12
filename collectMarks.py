# _*_ coding: utf_8  _*_
# read marks of many workbooks then write the marks togather for every student.
# it is too slowly. 2017-2-11  now it is quickly. 2017-2-12

import openpyxl
import os
import re
import logging
import pprint


STUDENT_COUNT = 38
DIRNAME = 'd:\\_PythonWorks\\excelOperate\\cj-2016201701'
CLASSREG = re.compile(r'\d{7}') # CLASSREG = re.compile(r'\d{7}[zZ]?')  # CLASSREG = re.compile('\d*7')
COLUMNS_MAP = {
        '课堂平时成绩': {'column': 'J', 'index': 3},
        '课堂期末成绩': {'column': 'M', 'index': 4},
        '课堂总成绩': {'column': 'O', 'index': 5},
        '实践成绩': {'column': 'Q', 'index': 6},
        '实验成绩': {'column': 'R', 'index': 7},
        '总成绩': {'column': 'S', 'index': 8},
        }
# or use tuple
# COLUMNS_MAP = (
#        (' 课堂平时成绩', 'J', 3),
#       ...
#       )

# logging.disable(logging.CRITICAL)
# logging.basicConfig( filename='loglearn.txt',level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )                   
logging.basicConfig( level = logging.ERROR, format = ' %(asctime)s - %(levelname)s - %(message)s' )
# logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.critical('--------Start of program---------')

studentNo = input('please input the students No: ')
if( studentNo == '' ):
    studentNo = '201510040101'
    print(' '*2+studentNo)    
logging.info('studentNo is:%s', studentNo)

className = studentNo[2:6]+studentNo[7:10]
logging.info('className is:%s',' '*2+className)

# The student's marks need to write in the cell of the excel
# List is reasonable at here.
studentMarks = [
    ['课程',],
    ['学号', ],
    ['姓名', ],
    ['课堂平时成绩',],
    ['课堂期末成绩',],
    ['课堂总成绩',],
    ['实践成绩',],
    ['实验成绩',],
    ['总成绩',],
    ]
# Read marks
for file in os.listdir(DIRNAME):
    logging.debug(file)
    classSearch = CLASSREG.search(file) # a = re.findall(CLASSREG,file)
    if not classSearch or classSearch.group() != className:
        logging.debug('Doesn\' match!')
        continue
    logging.debug(classSearch.group())
    logging.debug(className)
    fullname = DIRNAME + '\\' + file
    wb = openpyxl.load_workbook( fullname)
    sheet = wb.get_active_sheet()
    for twoDigit in range(STUDENT_COUNT):                   ####
        studentNum = str(int(studentNo) + twoDigit)    #####
        for row in range(1,sheet.max_row):                  #####
            logging.info('row is:%d',row)
            logging.info('学号:%s' % str(sheet['B'+str(row)].value))
            if sheet['B'+str(row)].value != int(studentNum):               #  学号在B列
                continue
            logging.debug( ' '*2+str(row)+'  '+sheet['D'+str(row)].value )       #  姓名在D列
            studentMarks[0].append(file)
            studentMarks[1].append(studentNum)
            studentMarks[2].append(sheet['D'+str(row)].value)
            for (k, v) in COLUMNS_MAP.items():
                logging.debug(k + str( sheet[v['column']+str(row)].value ) )  #  课堂平时成绩在J列
                studentMarks[v['index']].append(sheet[v['column']+str(row)].value)
            # 不确定是不是需要
            break
        # logging.error( 'error test')

# Write marks    
wb = openpyxl.Workbook()
wb.create_sheet(index=0,title=className)
sheet = wb.get_sheet_by_name(className)
col = 2       
for val in studentMarks:
    logging.info(val)
    for n in range(len(studentMarks[0])):         
        logging.info(val[n])
        sheet.cell(row = 2 + n,column = col).value = val[n]
    col +=1
newfullname = DIRNAME + '\\' + className + '.xlsx'
wb.save(newfullname)

print('Done!')
logging.critical('--------End---------')

"""
201510040101

"""
