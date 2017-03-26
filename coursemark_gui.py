#! python3
# _*_ coding: utf_8  _*_
'''list the ahead 8 students' mark.        2017-3-5
try to list the courses. then we can choose the course by click.  2017-3-6
try to arrange the widgets.                2017-3-7
try to add a Listbox.  it can work. but not strong. minus can not work. 2017-3-15
Minus mark can be written now.             2017-3-16  morning
It can be run arbitrary.                   2017-3-16  22:10
''' 

import openpyxl
import os
import re
from tkinter import *
import tkinter as tk
from tkinter import ttk

from getdir import *

# course
course_reg = re.compile(r'-([a-z]{3,11})-')     # 2017-3-4 debug.

performance_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['提出问题',],
    ['回答问题',],
    ['课堂作业',],
    ['作业1',],['作业2',],['作业3',],['作业4',],['作业5',],['作业6',],['作业7',],['是否已交课堂作业',],
    ]
lab_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],['操作5',],['操作6',],['操作7',],['操作8',],
    ['数据1',],['数据2',],['数据3',],['数据4',],['数据5',],['数据6',],['数据7',],['数据8',],   
    ['报告1',],['报告2',],['报告3',],['报告4',],
    ]
design_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['设计1',],['设计2',],['设计3',],['设计4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],  
    ['报告1',],['报告2',],
    ]
practice_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],   
    ['报告1',],['报告2',],    
    ]

tagdict = {'performance':performance_tag,
           'lab':lab_tag,
           'design':design_tag,
           'practice':practice_tag}

class Application(tk.Tk): # 继承自 tk.Tk
    '''界面、逻辑分离示例'''
    
    def __init__(self):
        '''初始化'''
        super().__init__() # 有点相当于tk.Tk()
        
        self.create_widgets()

    def create_widgets(self):
        'Create all kinds of widgets.'

        '''界面'''
        self.mainframe = ttk.Frame(self, padding="9 9 12 12") # 注意ttk.Frame()的第一个参数为self，因为这个类继承自tk.Tk类
        self.mainframe.grid(column=0, row=0, sticky=(tk.N, tk.W, tk.E, tk.S))
        self.mainframe.columnconfigure(0, weight=1)
        self.mainframe.rowconfigure(0, weight=1)

        # Create"列出课程" button .
        Button( self.mainframe, text = "列出课程", command=self.listfile).grid( column=0, row=0, sticky=(W))

        # Create"选择课程" button .
        Button( self.mainframe, text = "选择课程", command=self.sele_course).grid( column=0, row=3, sticky=(W))

        # "前8名:" button .
        Button(self.mainframe, text = "前8名:", command = self.ahead).grid( column=0, row=5, sticky=(W))

        # Create listbox.
        self.course =  Listbox(self.mainframe, width=85, height=6)
        self.course.grid( column=0, row=1, sticky=(W))

        # label.
        Label( self.mainframe, text = '所选课程').grid(  column=1, row=3, sticky=(W))

        self.contents = StringVar()
        self.contents.set('test.')        # set it to some value
        # Entry.
        self.course_ent = Entry( self.mainframe, width=85, textvariable = self.contents)
        self.course_ent.grid( column=0, row=4, sticky=(W))
        self.course_ent["textvariable"] = self.contents    # tell the entry widget to watch this variable
        self.course_ent.bind('<Key-Return>', self.markin)  # when the user hits return
        
        # text.
        self.ranktext = Text(self.mainframe, height=10, width=36, wrap='word')
        self.ranktext.grid( column=0, row=6, sticky=(W))

        # "加减分项目" button .
        Button(self.mainframe, text = "加减分项目", command = self.list_item).grid( column=0, row=2, sticky=(W))

        # "旷课者查询" button .
        Button(self.mainframe, text = "旷课者查询", command = self.find_absent).grid( column=1, row=0, sticky=(W))

        # "清除上次课堂作业上交记录" button .
        Button(self.mainframe, text = "清除上次课堂作业上交记录", command = self.clr_absent).grid( column=0, row=9, sticky=(W))

        # "作业未做者查询" button .
        Button(self.mainframe, text = "作业未做者查询", command = self.nohomework).grid( column=1, row=2, sticky=(W))

        # "退出"button.
        Button(self.mainframe, text="退出", fg="red", command=self.mainframe.destroy).grid(  column=0, row=11, sticky=(W))

    def sele_course(self):
        
        global fulllist, NUM
        if NUM >= 0:
            # every click, NUM decrease 1. to select the next course.
            NUM -= 1       #  decrease must be place here.
            self.contents.set( fulllist[NUM])
            if NUM < 0:
                NUM = len( fulllist) - 1
        pass

    def listfile(self):
        
        global fulllist
        self.course.delete( 0, END)
        for coursen in fulllist:
            self.course.insert( END, coursen)

    def list_item(self):
        
        global fulllist, NUM
        
        # list the item according to the coursetpye.
        self.course.delete( 0, END) # self.ranktext.delete(0.0, END)
        coursetype = course_reg.search( fulllist[NUM]).group(1)
        k = 0
        for val in tagdict[coursetype]:
            self.course.insert( END, str(k)+','+str(val)+'\n')
            k += 1
        self.contents.set('请输入项目代号2位学号3位及分数-2分记为12：0511102')
        pass

    def markin( self, event):    #    event.??    ok.   17-3-15.
        # modify the mark.
        
        global fulllist, NUM
        
        st = self.course_ent.get()
        print( st)
        if st.isdigit():
            itemnum = 6 + int( st[:2])

            wb = openpyxl.load_workbook(fulllist[NUM])
            sheet = wb.get_active_sheet()

            studnum = st[2:5]
            for row in range(3,sheet.max_row + 1):
                if str( sheet[ 'b'+str( row)].value)[-3:] == studnum:           #  学号在B列
                    # Write
                    if itemnum == 11:
                        sheet.cell(row=row,column=19).value += 1
                    dp = studnum +':'+ str( sheet.cell(row = row,column = itemnum).value)
                    dp += '总分:'+ str( sheet.cell(row = row,column = 4).value)
                    self.ranktext.insert( END, dp+' ')                       # 显示加之前的分数
                    mark = st[6:]
                    if st[5]=='0':
                        sheet.cell(row = row,column = itemnum).value += int(mark)        # 加上要加的分数
                        sheet.cell(row = row,column = 4).value += int(mark)              # 总分也加上该分数
                    elif st[5]=='1':
                        sheet.cell(row = row,column = itemnum).value -= int(mark)    # 减去要减的分数
                        sheet.cell(row = row,column = 4).value -= int(mark)          # 总分也减去该分数
                    dp = str( sheet.cell( row = row,column = itemnum).value)
                    dp += '总分:'+ str( sheet.cell(row = row,column = 4).value)
                    self.ranktext.insert( END, dp+' ')                     # 显示加之后的分数
                    break
                
            while True:
                try:    
                    wb.save(fulllist[NUM])
                except PermissionError:
                    input('Please close the workbook.')
                else:
                    break
            pass
    
    def ahead(self):
        
        global fulllist, NUM
        # Read the marks.
        wb = openpyxl.load_workbook( fulllist[NUM])
        sheet = wb.get_active_sheet()
        marks = []
        for row in range(3,sheet.max_row + 1):
            marks.append((sheet.cell(row = row,column = 4).value, sheet['b'+str(row)].value, sheet['c'+str(row)].value))
        wb.save( fulllist[NUM])
        
        self.ranktext.delete(1.0, END)
        self.ranktext.insert(END, '前8名为：\n')
        # rank the marks.
        marks.sort(reverse=True)
        for k in marks[:8]:
            # insert the ahead marks to text.
            self.ranktext.insert(END, str(k)+'\n')
        pass

    def find_absent( self):

        global fulllist, NUM
        # Open the book.
        col=19        # column 12 is "是否上交课堂作业"
        wb = openpyxl.load_workbook( fulllist[NUM])
        sheet = wb.get_active_sheet()
        ab_stu = []
        for row in range(3,sheet.max_row + 1):
            if not sheet.cell( row=row,column=col).value:
                ab_stu.append( sheet[ 'b'+str( row)].value)
        wb.save( fulllist[NUM])
        
        self.course.delete( 0, END)
        if len( ab_stu)<5:   # len( ab_stu) = 0 时，会插入一个空列，感觉比什么都没有更踏实。
            self.course.insert( 0, ab_stu)
        else:    #   len( ab_stu)>=5
            for k in range( len( ab_stu)//5):
                self.course.insert( 0, ab_stu[ 5*k:5*k+5])
            if len( ab_stu)%5:
                self.course.insert( 0, ab_stu[ 5*k+5:])
        pass

    def nohomework( self):

        global fulllist, NUM
        # Open the book.
        col=12        # column 12 is "作业1"
        wb = openpyxl.load_workbook( fulllist[NUM])
        sheet = wb.get_active_sheet()
        nohome = []
        for row in range(3,sheet.max_row + 1):
            if not sheet.cell( row=row,column=col).value:
                nohome.append( sheet[ 'b'+str( row)].value)
        wb.save( fulllist[NUM])
        
        self.course.delete( 0, END)

        if len( nohome) < 5:
            self.course.insert( 0, nohome)
        else:
            for k in range( len( nohome)//5):
                self.course.insert( 0, nohome[ 5*k:5*k+5])
            if len( nohome)%5:
                self.course.insert( 0, nohome[ 5*k+5:])
        pass

    def clr_absent(self):

        global fulllist, NUM
        # Open the book.
        wb = openpyxl.load_workbook( fulllist[NUM])
        sheet = wb.get_active_sheet()
        for row in range(3,sheet.max_row + 1):
            sheet.cell(row=row,column=19).value = 0
        wb.save( fulllist[NUM])
        pass

def getfile():

    global fulllist, NUM
    # Get the directory name.
    DIRNAME = getdir()
    # Get the filename list.
    FILELIST = os.listdir( DIRNAME)
    # Get the filename list include coursetype.
    filelist = filesele( FILELIST, course_reg)
    # sort the filelist. so the index of the file is nochange.
    filelist.sort()       
    # File full name list.
    fulllist = getfull( DIRNAME, filelist)
    NUM = len( fulllist) - 1

if __name__ == '__main__':  #  __main__ is not correct.


    getfile()
    
    # 实例化Application
    app = Application()
    
    # 设置窗口标题
    app.title("课程平时成绩")
    
    # 主消息循环:
    app.mainloop()
    
    print('End')

    
