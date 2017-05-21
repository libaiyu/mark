#! python3
# Put the final_mark into the excel file.

import os
import datetime
import openpyxl
from getdir import getdir, filesele, getfull, getbackup, getfile  # Get the directory name in this module.

itemnum = 22    # Final mark in column 22
studnum = []
mark = []

def backup( content):
    ''' Add time and write down some information( filename, input content, etc)
'''

    backupfile = getbackup()   # get backup() must run after getdir().
    memory_file = open( backupfile, 'a')
    memory_file.write( '\n'+content+'\n')
    memory_file.close()

def get_ex_file():
    ''' List all the excel files that record marks.

then select the file that will be used this time.
'''
    excel_file = getfile() # input( 'Please input the excel file that need write marks.')
    x = 0
    for e in excel_file:
        print( str(x), e)
        x += 1
    filenum = int( input('select file.'))
    backup( excel_file[ filenum])
    
    return excel_file[ filenum]

def get_f_mark():
    ''' Get the multi marks from input.

11189 means the last three digits of students' number is 111, mark is 89.

10288,10386,20591
'''
    st = input('Please input the marks split by ",":')   #  multi marks split by ",".
    backup( st)

    # split each mark.
    mm = st.split(',')
    # get the studnum, mark list.


    for e in range(len(mm)):
        if not mm[e].isdigit():    # each mark should be digit. if so, it will write in.
            break                  # till mark is not digit, break.

        if len(mm[e]) == 5:          # student number(3 digits), mark.
##            print(mm[e],type(mm[e]))
            studnum.append( mm[e][:3])
            class_temp = mm[e][0]
            marktemp = mm[e][3:]
            mark.append( marktemp)
        elif len(mm[e]) == 3:        # student number change(3 digits), mark no change.
            studnum.append( mm[e])
            class_temp = mm[e][0]
            mark.append( marktemp)   #  mark no change.
        elif len(mm[e]) == 4:          # student number(2 digits), mark.
            mmm = class_temp + mm[e][:2]
            studnum.append( mmm)
            marktemp = mm[e][2:]
            mark.append( marktemp)
        elif len(mm[e]) == 2:        # student number change(2 digits), mark no change.
            mmm = class_temp + mm[e]
            studnum.append( mmm)
            mark.append( marktemp)   #  mark no change.            
        else:
            print('数字位数不对。')
    pass
    print( studnum)
    print( mark)

    return None

def write_in( sele_file):
    ''' Write the final_mark in the excel file.
'''

    # open the excel file.
    wb = openpyxl.load_workbook( sele_file)
    sheet = wb.get_active_sheet()

    # write the marks.
    for e in range(len(studnum)):
        for row in range(3,sheet.max_row + 1):
            if str( sheet[ 'b'+str( row)].value)[-3:] == studnum[e]:           #  学号在B列
                # Write
                sheet.cell(row = row,column = itemnum).value = int(mark[e])    # 加上要加的分数
                break
    pass

    # save the excel file.
    while True:
        try:    
            wb.save( sele_file)
        except PermissionError:
            input('Please close the workbook.')
        else:
            break
    pass

def final_mark_in():
    ''' Put the final_mark into the excel file.
'''
    
    # Get the excel file.
    sele_file = get_ex_file()  # This must run before backup() run.
    
    # backup the detail time.
    ##  import datetime
    t = datetime.datetime.now()
    detail_time = '\n'+str( t.year)+'-'+str( t.month)+'-'+str( t.day)+','+str( t.hour)+':'+str( t.minute)+':'+str( t.second)
    backup( detail_time)    

    # Get the final mark.
    get_f_mark()

    # Write the final_mark into the open excel file.
    write_in( sele_file)
    
        
if __name__ == '__main__':
    final_mark_in()


