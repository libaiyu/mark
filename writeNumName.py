#! python 3
# _*_ coding: utf_8  _*_
# Read the number and name of the students from the file that include "学生名单".
# Then write the number and name of the students to the new files that will note the mark for students.
# find a bug. list of students need initial in the loop. not out of the loop.

import openpyxl
import os
import re
import logging

# logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.basicConfig( level = logging.ERROR, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.critical('--------Start of program---------')

# find the file that include '学生名单' in file
ChineseReg = re.compile(r'学生名单')
# find the class
classReg = re.compile(r'\d{7}(-\d)?')
# find the course
courseReg = re.compile(r'-(\w{3,11})-')

dirname = 'd:\\_PythonWorks\\excelOperate\\pscj161702'
for file in os.listdir(dirname):
    if ChineseReg.search(file):
        # read the number and name of the students from the file that include "学生名单"
        fullname = dirname + '\\' + file
        logging.info(ChineseReg.search(file).group())    # find the file that have "学生名单" in its name.    
        className = classReg.search(file).group()
        logging.info(className)        # find the class name in file name.
        # initial the list for a new className
        students = [
                    [ ],
                    [ ],
                    ['总分',],
                    ['初始分',],
                    ]
        wb = openpyxl.load_workbook(fullname)
        sheet = wb.get_active_sheet()
        # Read
        for row in range(1,sheet.max_row + 1):
            logging.debug(sheet['b'+str(row)].value)           #  学号在B列
            if sheet['b'+str(row)].value:                      #  学号在B列
                logging.debug(sheet['d'+str(row)].value)              #  姓名在d列
                students[0].append(str(sheet['B'+str(row)].value))    #  学号在B列 
                students[1].append(sheet['d'+str(row)].value)         #  姓名在D列
                students[2].append(65)         #  总分初始值为65
                students[3].append(65)         #  初始分设定为65

        # write the number and name of the students to the new files that will note the mark for students.
        count = 0
        for file in os.listdir(dirname):
            classReg2 = re.compile(className + '\.')
            if classReg2.search(file):
                logging.info(classReg2.search(file))
                if ChineseReg.search(file) == None:
                    fullname = dirname + '\\' + file
                    wb = openpyxl.load_workbook(fullname)
                    sheet = wb.get_active_sheet()
                    # Write
                    col = 2       
                    for val in students:
                        logging.info(val)
                        for n in range(len(students[0])):         
                            logging.info(val[n])
                            sheet.cell(row = 2 + n,column = col).value = val[n]
                        col +=1
                    wb.save(fullname)
                    count += 1
                    print('There are %d files for class: %s have been written!' % (count,className))

logging.critical('-------End--------')




