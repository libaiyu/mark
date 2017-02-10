# _*_ coding: utf_8  _*_
 
import logging
import os
import openpyxl
import pprint
import re

# logging.disable(logging.CRITICAL)
# logging.basicConfig( filename='loglearn.txt',level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )                   
logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.critical('--------Start of program---------')

# Write marks
wb = openpyxl.Workbook()
wb.create_sheet(index=0,title='example_2')
sheet = wb.get_sheet_by_name('example_2')
# The student's marks need to write in the cell of the excel
# List is reasonable at here.
studentsMarks = [
    ['学号', '201510040101','201510040102','201510040103',],
    ['课堂平时成绩',85,95,86,],
    ['课堂期末成绩',76,86,78,],
    ['课堂总成绩',76,87,80,],
    ['实践成绩',76,75,85,],
    ['实验成绩',76,86,90,],
    ['总成绩',76,90,89,],
    ]
for k in range(len(studentsMarks)):
    studentsMarks[k].append('78')
col = 2       
for val in studentsMarks:
    logging.info(val)
    for n in range(len(studentsMarks[0])):         
        logging.info(val[n])
        sheet.cell(row = 2 + n,column = col).value = val[n]
    col +=1

wb.save('example_2.xlsx')
 


