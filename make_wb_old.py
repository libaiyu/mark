#! python 3
# _*_ coding: utf_8  _*_


import openpyxl
import os
import re

from getdir import getdir    # 2017-3-22,  2017-5-13


regex = re.compile('(\w+)-(\d+)(-\d)?|(\w+)-([a-z]+)-(\d+)(-\d)?')   #  使用分组匹配
'''
>>> s = '模拟电子技术-performance-1523701-test'
>>> import re;regex = re.compile('(\w+)-(\d+)(-\d)?|(\w+)-([a-z]+)-(\d+)(-\d)?')
>>> regex.match( s)
<_sre.SRE_Match object; span=(0, 26), match='模拟电子技术-performance-1523701'>
>>> regex.match( s).groups()
(None, None, None, '模拟电子技术', 'performance', '1523701', None)
>>> import re
>>> s = '学生名单-1620604-5'
>>> regex = re.compile('(\w+)-(\d+)(-\d)?|(\w+)-([a-z]+)-(\d+)(-\d)?')
>>> regex.match( s)
<_sre.SRE_Match object; span=(0, 14), match='学生名单-1620604-5'>
>>> regex.match( s).groups()
('学生名单', '1620604', '-5', None, None, None, None)
'''
course_reg = re.compile(r'-([a-z]{3,11})-')

performance_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['提出问题',],
    ['回答问题',],
    ['课堂作业',],
    ['作业1',],['作业2',],['作业3',],['作业4',],['作业5',],['作业6',],['作业7',],['是否已交课堂作业',],
    ]
lab_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],['操作5',],['操作6',],['操作7',],['操作8',],
    ['数据1',],['数据2',],['数据3',],['数据4',],['数据5',],['数据6',],['数据7',],['数据8',],
    ['报告1',],['报告2',],['报告3',],['报告4',],    
    ]
design_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['设计1',],['设计2',],['设计3',],['设计4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],
    ['报告1',],['报告2',],    
    ]
practice_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],
    ['报告1',],['报告2',],    
    ]
tagdict = {'performance':performance_tag,
           'lab':lab_tag,
           'design':design_tag,
           'practice':practice_tag}


def read_num_name( file):
    ''' Read the number and name of the students from the file that include "学生名单".
'''

    wb = openpyxl.load_workbook( file)
    sheet = wb.get_active_sheet()
    # Read the stud_num and name.
    stud_num = []
    stud_name = []
    for row in range(1,sheet.max_row + 1):        
        if sheet['b'+str(row)].value:                         #  学号在B列
            stud_num.append( str(sheet['B'+str(row)].value))  #  学号在B列 
            stud_name.append( str(sheet['D'+str(row)].value))  #  姓名在D列
    return stud_num, stud_name

def write_num_name( file, stud_num, stud_name):
    ''' Write the number and name of the students to the new files that will note the mark for students.
'''           

    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    # Write the stud_num and name.
    r = 2
    c = 2
    for each_num, each_name in zip( stud_num, stud_name):        
        sheet.cell( row = r,column = c).value = each_num
        sheet.cell( row = r,column = c+1).value = each_name
        r += 1
    wb.save( file)


def add_num_name( file):
    ''' Read the number and name of the students from the file that include "学生名单".
Then write the number and name of the students to the new files that will note the mark for students.
'''
    pass
##read_num_name( file)
##
##
##write_num_name()

'''

    
    for file in os.listdir(dirname):
        print( file)
        # input('debug')
        if regex.match( file) is not None:      # match the first case. (\w+)-(\d+)(-\d)?
            matchgroup = regex.match( file).groups()
            matchlist.append( file)      #  get the match file list.
        if matchgroup[0]:    #  find the file that have "学生名单" in its name. 
            logging.info( matchgroup[0])    #  显示 "学生名单"
            
            # read the number and name of the students from the file that include "学生名单"
            fullname = dirname + '\\' + file
              
            if matchgroup[2]:
                className = matchgroup[1] + matchgroup[2]
            else:
                className = matchgroup[1]

      read_num_name()


            
    print()
    # write the number and name of the students to the new files that will note the mark for students.
    count = 0
    for file in matchlist:
        matchgrp = regex.match( file).groups()
        if matchgrp[6] is not None:    # get the class name in the file that include course.
            className = matchgrp[5] + matchgrp[6]
        elif matchgrp[5] is not None:
            className = matchgrp[5]
        else:
            className = None
        logging.info( className)
        if className is not None:
            fullname = dirname + '\\' + file

            write_num_name()

            count += 1
            print('%s have been added Number and name!\nFile %d .' % (file, count))
'''

def test():
    '''  '''
    dirname = 'd:\\_PythonWorks\\Opexcel\\pscj171801'
    os.chdir( dirname)
    matchlist = []
    for file in os.listdir( '.'):
        print( file)
##        input('debug')
        if regex.match( file) is not None:      # match the first case. (\w+)-(\d+)(-\d)?
            matchgroup = regex.match( file).groups()
            matchlist.append( file)      #  get the match file list.
        if matchgroup[0]:    #  find the file that have "学生名单" in its name.
            num, name = read_num_name( file)
##            print( num_name)
            newfile = '模拟电子技术_performance_' + file
            write_num_name( newfile, num, name)
            newfile = '模拟电子技术_lab_' + file
            write_num_name( newfile, num, name)
            newfile = '模拟电子技术_design_' + file
            write_num_name( newfile, num, name)


if __name__ == '__main__':

    print('---------Begin--------')
    test()
##    add_num_name()
    print('---------End--------')

            
