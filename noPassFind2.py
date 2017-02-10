# _*_ coding: utf_8  _*_
 
import openpyxl
import os
import re
import logging

classReg = re.compile(r'\d{7}')

logging.disable(logging.CRITICAL)
# logging.basicConfig( filename='loglearn.txt',level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )                   
logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.critical('--------Start of program---------')


# Read marks

for fileName in os.listdir('d:\\_PythonWorks\\execlOperate\\working'):
    logging.debug(fileName)
    # The student's marks need to write in the cell of the excel
    # List is reasonable at here.
    studentMarks = [
        ['课程',],
        ['学号', ],
        ['姓名', ],
        ['课堂平时成绩',],
        ['课堂期末成绩',],
        ['课堂总成绩',],
        ['实践成绩',],
        ['实验成绩',],
        ['总成绩',],
        ]
    if classReg.search(fileName): 
        wb = openpyxl.load_workbook( fileName ,'r' )
        sheet = wb.get_active_sheet()
        for row in range(3,150):
            logging.info('row is:%d',row)
            logging.info( str( sheet['J'+str(row)].value ) )
            if sheet['J'+str(row)].value:
                if( int(sheet['J'+str(row)].value) < 59 ):
                    logging.debug( '课程'+sheet['B'+str(row)].value )                    #  课程在B列
                    logging.debug( '学号'+sheet['C'+str(row)].value )                    #  学号在C列
                    logging.debug( '姓名'+sheet['D'+str(row)].value )                    #  姓名在D列
                    logging.debug(' 课堂平时成绩：'+ str( sheet['E'+str(row)].value ) )  #  课堂平时成绩在E列
                    logging.debug(' 课堂期末成绩：'+ str( sheet['F'+str(row)].value ) )  #  课堂期末成绩在F列
                    logging.debug(' 课堂总成绩：'+ str( sheet['G'+str(row)].value ) )    #  课堂总成绩在G列
                    logging.debug(' 实践成绩：'+ str( sheet['H'+str(row)].value ) )      #  实践成绩在H列
                    logging.debug(' 实验成绩：'+ str( sheet['I'+str(row)].value ) )      #  实验成绩在I列
                    logging.debug(' 总成绩：'+ str( sheet['J'+str(row)].value ) )        #  总成绩在J列
                    studentMarks[0].append(sheet['B'+str(row)].value)
                    studentMarks[1].append(sheet['C'+str(row)].value)
                    studentMarks[2].append(sheet['D'+str(row)].value)
                    studentMarks[3].append(sheet['E'+str(row)].value)
                    studentMarks[4].append(sheet['F'+str(row)].value)
                    studentMarks[5].append(sheet['G'+str(row)].value)
                    studentMarks[6].append(sheet['H'+str(row)].value)
                    studentMarks[7].append(sheet['I'+str(row)].value)
                    studentMarks[8].append(sheet['J'+str(row)].value)                  

        # Write marks    
        wb = openpyxl.Workbook()
        wb.create_sheet(index=0,title=fileName)
        sheet = wb.get_sheet_by_name(fileName)
        sheet.column_dimensions['a'].width = 2
        sheet.column_dimensions['b'].width = 30
        sheet.column_dimensions['c'].width = 13
        letters = ['d','e','f','g','h','i','j']
        for colu in letters:
            sheet.column_dimensions[colu].width = 10        
        col = 2       
        for val in studentMarks:
            logging.info(val)
            for n in range(len(studentMarks[0])):         
                logging.info(val[n])
                sheet.cell(row = 2 + n,column = col).value = val[n]
            col +=1
        newExcelName = 'BK_' + fileName
        wb.save(newExcelName)
        print('one file is written!')
print('Done!')


