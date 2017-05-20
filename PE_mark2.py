#! python 3
# _*_ coding: utf_8  _*_

'''

'''

import openpyxl
import os
import re
import logging


MAXROW = 106
Reg = re.compile( r'标准表')

##TYPE_AGE = ['甲', '乙', '丙', '丁']
China_str = ['100米','跳远','铅球','８００米','男子成绩','女子成绩']
PE_items = ['100米','跳远','铅球','８００米']
TYPE = ['男子成绩','女子成绩']

measure_v = []

def get_params():
    '''先获取所需的信息：姓名，男女，项目，测量值，组别。

'''
    wb = openpyxl.load_workbook('高考体育成绩测量表.xlsx')
    sheet = wb.get_active_sheet()
    col = 2
    for row in range(1,sheet.max_row+1):
        val = sheet.cell( row=row, column=col).value
        if val:
            m_val = []
            type_age = sheet.cell( row=row, column=col-1).value
            for k in range(4):
                m_val.append( [ PE_items[k], sheet.cell( row=row, column=col+k+2).value])
            measure_v.append( ( type_age, val, sheet.cell( row=row, column=col+1).value, m_val))
    for t in range(3):
        print( measure_v[ t])
    print()

def get_mark( type_age, gender, item_name, measure_val):
    '''然后根据参数中提供的信息找到相应的成绩。
'''
    
    if gender:
        mark = mark_woman[(type_age,PE_items[item_name])][measure_val]
    else:
        mark = mark_man[(type_age,PE_items[item_name])][measure_val]
##    print(mark)
    return mark

# logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.basicConfig( level = logging.ERROR, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.critical('--------Start of program---------')

item_head = {}
mark_man = {}
mark_woman = {}
wb_val = {}

for e in os.listdir():

    if Reg.search( e):
        print( e)
        type_age = e[3]

        '''读成绩对照表，生成字典的字典。以供后续查询。
        '''
        wb = openpyxl.load_workbook( e)
        sheet = wb.get_active_sheet()
        for col in range(1,sheet.max_column+1):
            key1 = col
            col_dict = {}
            for row in range(1,sheet.max_row+1):
                key2 = row
                val = sheet.cell( row=row, column=col).value # 给定的列与行的值。
        ##        print( type( val))
        ##        input()
        ##        if val is not None:
        ##            col_dict[ key2] = val
                col_dict[ key2] = val   #  字典－－遍历所有行。　｛行：值｝。
            wb_val[ key1] = col_dict    #  字典－－遍历所有列。　｛列：｛行：值｝｝。
        print( key1, key2, val)

        for col, row_val in wb_val.items():
            for row, val in row_val.items():
                if val in China_str:
                    if val in PE_items:   #  如果值是项目名。
                        item_head[val] = ( col, row)   #  字典--项目名为键，所在的（列，行）为值。
        print(item_head)
        print()

        for val, ( col, row) in item_head.items():
            key1 = ( type_age, val)    #  组别，项目名
            item_man = {}
            item_woman = {}
            for r in range( row+2, MAXROW):
                key_m = wb_val[col][r]        #  男　的键　为　测量值
                val_m = wb_val[col+1][r]      #  男　的值　为　成绩
                item_man[key_m] = val_m       #  男　的字典　　测量值：成绩
                key_wo = wb_val[col+2][r]     #  女　的键　为　测量值
                val_wo = wb_val[col+3][r]     #  女　的值　为　成绩
                item_woman[key_wo] = val_wo   #  女　的字典
            mark_man[key1] = item_man         #  男　的字典的字典　项目名　为　键，（测量值：成绩）　为　值。
            mark_woman[key1] = item_woman     #  女　的字典的字典　项目名　为　键，（测量值：成绩）　为　值。

##print('男子成绩\n',mark_man,'\n')
##print('女子成绩\n',mark_woman,'\n')

'''读　‘高考体育成绩测量表.xlsx’，生成　字典的字典
'''
get_params()

for each in measure_v:
    type_age = each[0]
##    print( each)
    if each[2] == '男':
        gender = 0
    elif each[2] == '女':
        gender = 1
    for k in range(4):
        item_name = k
        measure_val = each[3][k][1]
        mark = get_mark( type_age, gender, item_name, measure_val)
        each[3][k].append( mark)
    print( each)
##    input('for debug')
##for each in measure_v:
##    print( each)
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
row = 0
for each in measure_v:
    row += 1
    sheet.cell( row = row, column = 1).value = each[1]
    sheet.cell( row = row, column = 2).value = each[2]
    sheet.cell( row = row, column = 3).value = each[3][0][0]
    sheet.cell( row = row, column = 4).value = each[3][0][1]
    sheet.cell( row = row, column = 5).value = each[3][0][2]
    sheet.cell( row = row, column = 6).value = each[3][1][0]
    sheet.cell( row = row, column = 7).value = each[3][1][1]
    sheet.cell( row = row, column = 8).value = each[3][1][2]
    sheet.cell( row = row, column = 9).value = each[3][2][0]
    sheet.cell( row = row, column = 10).value = each[3][2][1]
    sheet.cell( row = row, column = 11).value = each[3][2][2]    
    sheet.cell( row = row, column = 12).value = each[3][3][0]
    sheet.cell( row = row, column = 13).value = each[3][3][1]
    sheet.cell( row = row, column = 14).value = each[3][3][2]
wb.save('体育成绩表.xlsx')

logging.critical('-------End--------')






