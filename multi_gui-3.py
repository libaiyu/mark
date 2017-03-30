
#! python3
# _*_ coding: utf_8  _*_
'''list the ahead 8 students' mark.        2017-3-5
try to list the courses. then we can choose the course by click.  2017-3-6
try to arrange the widgets.                2017-3-7
try to add a Listbox.  it can work. but not strong. minus can not work. 2017-3-15
Minus mark can be written now.             2017-3-16  morning
It can be run arbitrary.                   2017-3-16  22:10

多页面                2017-3-26
增加查询作业上交情况前的输入提示    2017-3-28
''' 

import openpyxl
import os
import re
from tkinter import *
import tkinter as tk
from tkinter import ttk

from getdir import *

# course
course_reg = re.compile(r'-([a-z]{3,11})-')     # 2017-3-4 debug.

global contents             # Public variable

performance_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['提出问题',],
    ['回答问题',],
    ['课堂作业',],
    ['作业1',],['作业2',],['作业3',],['作业4',],['作业5',],['作业6',],['作业7',],['是否已交课堂作业',],
    ]
lab_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],['操作5',],['操作6',],['操作7',],['操作8',],
    ['数据1',],['数据2',],['数据3',],['数据4',],['数据5',],['数据6',],['数据7',],['数据8',],   
    ['报告1',],['报告2',],['报告3',],['报告4',],
    ]
design_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['设计1',],['设计2',],['设计3',],['设计4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],  
    ['报告1',],['报告2',],
    ]
practice_tag = [
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],   
    ['报告1',],['报告2',],    
    ]

tagdict = {'performance':performance_tag,
           'lab':lab_tag,
           'design':design_tag,
           'practice':practice_tag}




import matplotlib
matplotlib.use("TkAgg")

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2TkAgg
from matplotlib.figure import Figure

import tkinter as tk
from tkinter import ttk


LARGE_FONT= ("Verdana", 12)



class Application(tk.Tk):
    '''
    平时成绩管理-多页面版
        界面与逻辑分离
    '''
    def __init__(self):
        
        super().__init__()

        self.iconbitmap(default="whoami.ico")
        self.wm_title("平时成绩管理-多页面版")
        
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand = True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}
        for F in (StartPage, PageOne, PageTwo, PageThree):
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")  # 四个页面的位置都是 grid(row=0, column=0), 位置重叠！！



        self.show_frame(StartPage)

        
    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise() # 切换，提升当前 tk.Frame z轴顺序（使可见）！！此语句是本程序的点睛之处

        
class StartPage(tk.Frame):
    '''选课程'''
    def __init__(self, parent, root):
        super().__init__(parent)
        label = tk.Label(self, text="选课程", font=LARGE_FONT)
        label.pack()

        button1 = ttk.Button(self, text="去到课堂", command=lambda: root.show_frame(PageOne)).pack()
        button2 = ttk.Button(self, text="去到作业", command=lambda: root.show_frame(PageTwo)).pack()
##        button3 = ttk.Button(self, text="去到课后", command=lambda: root.show_frame(PageThree)).pack()


        self.create_widgets()

    def create_widgets(self):
        'Create all kinds of widgets.'

        '''界面'''
        self.mainframe = ttk.Frame(self, padding="9 9 12 12") # 注意ttk.Frame()的第一个参数为self，因为这个类继承自tk.Tk类
        self.mainframe.pack()
        self.mainframe.columnconfigure(0, weight=1)
        self.mainframe.rowconfigure(0, weight=1)

        # Create"列出课程" button .
        Button( self.mainframe, text = "列出课程", command=self.listfile).pack()

        # Create listbox.
        self.course =  Listbox(self.mainframe, width=85, height=5)
        self.course.pack()

        # Create"选择课程" button .
        Button( self.mainframe, text = "选择课程", command=self.sele_course).pack()
        
        # label.
        Label( self.mainframe, text = '所选课程').pack()
        
        global contents             # Public variable
        contents = StringVar()
        contents.set('用于显示所选择的课程.每页都相同。')        # set it to some value
        
        # Entry.
        self.course_ent = Entry( self.mainframe, width=85, textvariable = contents)
        self.course_ent.pack()


    def listfile(self):     #  "列出课程"
        
        global fulllist
        self.course.delete( 0, END)
        for coursen in fulllist:
            self.course.insert( END, coursen)

    def sele_course(self):    # "选择课程"
        
        global fulllist, NUM
        if NUM <= 0:
            NUM = len( fulllist)
        if NUM > 0:
            # every click, NUM decrease 1. to select the next course.
            NUM -= 1       #  decrease must be place here.
            contents.set( fulllist[NUM])

        pass


class PageOne(tk.Frame):
    '''课堂'''
    def __init__(self, parent, root):
        super().__init__(parent)
        label = tk.Label(self, text="这是课堂", font=LARGE_FONT)
        label.pack()

        button1 = ttk.Button(self, text="回到选课程", command=lambda: root.show_frame(StartPage)).pack()
        button2 = ttk.Button(self, text="去到作业", command=lambda: root.show_frame(PageTwo)).pack()


        self.create_widgets()

    def create_widgets(self):
        'Create all kinds of widgets.'

        '''界面'''
        self.mainframe = ttk.Frame(self, padding="9 9 12 12") # 注意ttk.Frame()的第一个参数为self，因为这个类继承自tk.Tk类
        self.mainframe.pack()
        self.mainframe.columnconfigure(0, weight=1)
        self.mainframe.rowconfigure(0, weight=1)

        # label.
        Label( self.mainframe, text = '所选课程').pack()
        
        # Entry 1.
        self.course_ = Entry( self.mainframe, width=85, textvariable = contents)
        self.course_.pack()

        # "前8名:" button .
        Button(self.mainframe, text = "前8名:", command = self.ahead).pack()
        
        # text.
        self.ranktext = Text(self.mainframe, height=12, width=65, wrap='word')
        self.ranktext.pack()

        # "清除上次课堂作业上交记录" button .
        Button(self.mainframe, text = "清除上次课堂作业上交记录", command = self.clr_absent).pack()

        # "加减分项目" button .
        Button(self.mainframe, text = "加减分项目", command = self.list_item).pack()
        
        self.cont = StringVar()
        self.cont.set('请输入项目代号2位学号3位及分数-2分记为12：0511102')   # set it to some value
        # Entry 2.
        self.course_ent = Entry( self.mainframe, width=85, textvariable = self.cont)
        self.course_ent.pack()
        self.course_ent["textvariable"] = self.cont    # tell the entry widget to watch this variable
        self.course_ent.bind('<Key-Return>', self.markin)  # when the user hits return

        # "旷课者查询" button .
        Button(self.mainframe, text = "旷课者查询", command = self.find_absent).pack()

        # Create listbox.
        self.course =  Listbox(self.mainframe, width=85, height=5)
        self.course.pack()

    def list_item(self):
        
        global fulllist, NUM
        
        # list the item according to the coursetpye.
        self.course.delete( 0, END) # self.ranktext.delete(0.0, END)
        coursetype = course_reg.search( fulllist[NUM]).group(1)
        k = 0
        for val in tagdict[coursetype][:6]:
            self.course.insert( END, str(k)+','+str(val)+'\n')
            k += 1
        self.cont.set('0411102,11203,21112,0512301,0521301,222,223')
        pass
    
    
    def markin( self, event):    #    event.??    ok.   17-3-15.
        # mark update.
        
        global fulllist, NUM

        # get the multi marks from Entry 2.
        st = self.course_ent.get()   #  multi marks split by ",".
        print( st)

        # split each mark.
        mm = st.split(',')       
        print (mm)

        # get the itemnum, studnum, mark list.
        itemnum = []
        studnum = []
        mark = []
        
        for e in range(len(mm)):
            if not mm[e].isdigit():    # each mark should be digit. if so, it will write in.
                break                  # till mark is not digit, break.

            if len(mm[e]) == 7:          # item, student number, mark.
                print(mm[e],type(mm[e]))
                ittemp = 6 + int( mm[e][:2])
                itemnum.append( ittemp)
                studnum.append( mm[e][2:5])
                marktemp = mm[e][5:]
                mark.append( marktemp)
            elif len(mm[e]) == 5:        # item no change, student number change, mark change.
                itemnum.append( ittemp)   # item no change.
                studnum.append( mm[e][:3])
                marktemp =  mm[e][3:]
                mark.append( marktemp)
            elif len(mm[e]) == 3:        # item no change, student number change, mark no change.
                itemnum.append( ittemp)
                studnum.append( mm[e])
                mark.append( marktemp)   #  mark no change.
            else:
                print('数字位数不对。')
                pass
        print( itemnum, studnum, mark)

        # open the excel file.
        wb = openpyxl.load_workbook(fulllist[NUM])
        sheet = wb.get_active_sheet()
        
        for e in range(len(studnum)):
            for row in range(3,sheet.max_row + 1):
                if str( sheet[ 'b'+str( row)].value)[-3:] == studnum[e]:           #  学号在B列
                    # Write
                    if itemnum[e] == 11:      # 课堂作业已做，说明来上课了。
                        sheet.cell(row=row,column=19).value += 1   # 上课记录，0表示缺课,1 is nomal, 2 means repeat write.
                    dp = studnum[e] +':'+ str( sheet.cell(row = row,column = itemnum[e]).value)
                    dp += '-'+ str( sheet.cell(row = row,column = 4).value)+'->'
##                    self.ranktext.insert( 0.0, dp+' ')                       # 显示加之前的分数
                    if mark[e][:1]=='0':
                        sheet.cell(row = row,column = itemnum[e]).value += int(mark[e][1:])        # 加上要加的分数
                        sheet.cell(row = row,column = 4).value += int(mark[e][1:])              # 总分也加上该分数
                    elif mark[e][:1]=='1':
                        sheet.cell(row = row,column = itemnum[e]).value -= int(mark[e][1:])    # 减去要减的分数
                        sheet.cell(row = row,column = 4).value -= int(mark[e][1:])          # 总分也减去该分数
                    dp += str( sheet.cell( row = row,column = itemnum[e]).value)
                    dp += '-'+ str( sheet.cell(row = row,column = 4).value)
                    self.ranktext.insert( 0.0, dp+'  ')                     # 显示加之后的分数   0.0  END
                    break
                
        while True:
            try:    
                wb.save(fulllist[NUM])
            except PermissionError:
                input('Please close the workbook.')
            else:
                break
        pass

    
    def ahead(self):
        
        global fulllist, NUM
        # Read the marks.
        wb = openpyxl.load_workbook( fulllist[NUM])
        sheet = wb.get_active_sheet()
        marks = []
        for row in range(3,sheet.max_row + 1):
            marks.append((sheet.cell(row = row,column = 4).value, sheet['b'+str(row)].value, sheet['c'+str(row)].value))
        wb.save( fulllist[NUM])
        
        self.ranktext.delete(1.0, END)
        self.ranktext.insert(END, '前8名为：\n')
        # rank the marks.
        marks.sort(reverse=True)
        for k in marks[:8]:
            # insert the ahead marks to text.
            self.ranktext.insert(END, str(k)+'\n')
        pass

    def find_absent( self):

        global fulllist, NUM
        # Open the book.
        col=19        # column 12 is "是否上交课堂作业"
        wb = openpyxl.load_workbook( fulllist[NUM])
        sheet = wb.get_active_sheet()
        ab_stu = []
        for row in range(3,sheet.max_row + 1):
            if not sheet.cell( row=row,column=col).value:
                ab_stu.append( sheet[ 'b'+str( row)].value)
        wb.save( fulllist[NUM])
        
        self.course.delete( 0, END)
        if len( ab_stu)<5:   # len( ab_stu) = 0 时，会插入一个空列，感觉比什么都没有更踏实。
            self.course.insert( 0, ab_stu)
        else:    #   len( ab_stu)>=5
            for k in range( len( ab_stu)//5):
                self.course.insert( 0, ab_stu[ 5*k:5*k+5])
            if len( ab_stu)%5:
                self.course.insert( 0, ab_stu[ 5*k+5:])
        pass


    def clr_absent(self):

        global fulllist, NUM
        # Open the book.
        wb = openpyxl.load_workbook( fulllist[NUM])
        sheet = wb.get_active_sheet()
        for row in range(3,sheet.max_row + 1):
            sheet.cell(row=row,column=19).value = 0
        wb.save( fulllist[NUM])
        pass


class PageTwo(tk.Frame):
    '''作业'''
    def __init__(self, parent, root):
        super().__init__(parent)
        label = tk.Label(self, text="这是作业", font=LARGE_FONT)
        label.pack()

        button1 = ttk.Button(self, text="回到选课程", command=lambda: root.show_frame(StartPage)).pack()
        button2 = ttk.Button(self, text="回到课堂", command=lambda: root.show_frame(PageOne)).pack()


        self.create_widgets()

    def create_widgets(self):
        'Create all kinds of widgets.'

        '''界面'''
        self.mainframe = ttk.Frame(self, padding="9 9 12 12") # 注意ttk.Frame()的第一个参数为self，因为这个类继承自tk.Tk类
        self.mainframe.pack()
        self.mainframe.columnconfigure(0, weight=1)
        self.mainframe.rowconfigure(0, weight=1)

        # label.
        Label( self.mainframe, text = '所选课程').pack()

        # Entry.
        self.course_ent = Entry( self.mainframe, width=85, textvariable = contents)
        self.course_ent.pack()

        # "提示输入第几次作业" label .
        Label(self.mainframe, text = "请输入要查询第几次作业").pack()

        self.homework = StringVar()
        self.homework.set('2')   # 2 means the second homework.
        # Entry 3.
        self.homew_ent = Entry( self.mainframe, width=3, textvariable = self.homework)
        self.homew_ent.pack()
        
        # "作业未做者查询" button .
        Button(self.mainframe, text = "作业未做者查询", command = self.nohomework).pack()
        
        # Create listbox.
        self.course =  Listbox(self.mainframe, width=85, height=5)
        self.course.pack()


    def nohomework( self):

        global fulllist, NUM

        self.course.delete( 0, END)
        # input('please input a number for the homework you want to check.')  # how to prompt .
        which = self.homew_ent.get()
##        self.cont.set["textvariable"] = '第几次作业？which ='
##        which = int( input( '第几次作业？'))
        self.course.insert( 0, which)
        col= int(which) + 11        # column 12 is "作业1"
        # Open the book.
        wb = openpyxl.load_workbook( fulllist[NUM])
        sheet = wb.get_active_sheet()
        nohome = []
        for row in range(3,sheet.max_row + 1):
            if not sheet.cell( row=row,column=col).value:
                nohome.append( sheet[ 'b'+str( row)].value)
##        try:
        wb.save( fulllist[NUM])
##        except KeyError:
##            # print('please input a number for the homework you want to check.')
##            self.course.set('please input a number for the homework you want to check.')
            
        if len( nohome) < 5:
            self.course.insert( 0, nohome)
        else:
            for k in range( len( nohome)//5):
                self.course.insert( 0, nohome[ 5*k:5*k+5])
            if len( nohome)%5:
                self.course.insert( 0, nohome[ 5*k+5:])
        pass


class PageThree(tk.Frame):
    '''课后'''
    def __init__(self, parent, root):
        super().__init__(parent)
        tk.Label(self, text="这是课后", font=LARGE_FONT).pack()

        button1 = ttk.Button(self, text="回到选课程", command=lambda: root.show_frame(StartPage)).pack()



def getfile():

    global fulllist, NUM
    # Get the directory name.
    DIRNAME = getdir()
    # Get the filename list.
    FILELIST = os.listdir( DIRNAME)
    # Get the filename list include coursetype.
    filelist = filesele( FILELIST, course_reg)
    # sort the filelist. so the index of the file is nochange.
    filelist.sort()       
    # File full name list.
    fulllist = getfull( DIRNAME, filelist)
    NUM = len( fulllist)

if __name__ == '__main__':  #  __main__ is not correct.


    getfile()
    
    # 实例化Application
    app = Application()
    
    # 主消息循环:
    app.mainloop()
    
    print('End')

