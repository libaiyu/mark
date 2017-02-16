#! python3
# _*_ coding: utf_8  _*_
# copy marks in many workbooks and put them togather for every student.

import openpyxl
import os
import re
import logging

# logging.basicConfig( level = logging.ERROR, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )

class collectmarks():
    """    collect marks of the students in several courses.    
    """

    def __init__(self):
        # the list for storing the classes and courses
        self.classlist = []
        # the row as the title 
        self.titlerow = []
        # the list of all rows of all workbooks. 
        self.allrow = []
        # the dictionary of the title and its index
        self.titledict = {}
        # after deleting null cell, we get the new title. 
        self.newtitle = []
        # according the index keep the nonnull cell. new all rows of all workbooks 
        self.newallrow = []
        # newtitle for dictionary
        self.newtitled = []
        # studentnumber as the key, course and marks form the dictionary as the value {studentnum: (course,studentname,[marks]}}
        self.studentcoursesm = {}
        
        self.coursemark = []
        self.classmark = {}
        self.studentm = []        

    # find out the classes and courses in the source directory
    def findclass(self, srcdir, relist):
        for file in os.listdir(srcdir):
            logging.info(file)
            CLASSREG = re.compile(relist)
            if CLASSREG.search(file):
                # tuple as a element of the list
                self.classlist.append((CLASSREG.search(file).group(1),file))     
        print(self.classlist)

    # find the title in anyone workbook
    def findtitle(self, srcdir):
        twoline = 0
        file = self.classlist[0][1]            
        fullname = srcdir + '\\' + file
        logging.info(fullname)         
        wb = openpyxl.load_workbook( fullname)
        sheet = wb.get_active_sheet()
        for row in range(1,sheet.max_row):
            # rowmark: store marks of one row
            # the first element of every row is file name
            rowmark = [file]       
            for col in range(1,sheet.max_column):
                rowmark.append(sheet.cell(row = row, column = col).value)
            # get the title line. it has two lines.
            while(twoline == 1):
                self.titlerow.append(rowmark)
                twoline += 1
            if rowmark[2] == '学    号' and twoline == 0:
                twoline += 1
                self.titlerow.append(rowmark)
        print(self.titlerow)

    # copy every row in all workbooks to form a list        
    def copyrows(self, srcdir):
        self.allrow = []   # for store all row marks        
        for t in self.classlist:
            file = t[1]            
            fullname = srcdir + '\\' + file
            logging.info(fullname)
            # copy row marks ,add course name            
            wb = openpyxl.load_workbook( fullname)
            sheet = wb.get_active_sheet()
            for row in range(1,sheet.max_row):
                # rowmark: store marks of one row
                # the first element of every row is file name
                rowmark = [file]       
                for col in range(1,sheet.max_column):
                    rowmark.append(sheet.cell(row = row, column = col).value)
                # add every row marks to form all row marks
                self.allrow.append(rowmark)

    # find the position of nonnull cells in title, 
    def findindex(self):
        firstline = ['课程名', '序号', '学    号', None, '姓  名', None, '班  级', None, None, None,
                     '课堂', None,    '课堂',      '课堂',   None,  '课堂',   None, '实践成绩', '实验成绩',
                     '总成绩', '特殊原因', None, '录入状态', '备  注', None]
        secondline = ['课程名', None,    None,      None,    None,  None,   None,   None, None, None,
                      '平时成绩', None,'期中成绩',  '期末成绩', None, '总成绩', None,   None,      None,
                      None,     None,     None,    None,      None,   None]
        self.titledict = {}
        self.titledict[0] = firstline[0]
        for pos in range(1,len(firstline)):
#            print(pos, firstline[pos], secondline[pos])
            if firstline[pos] != None and secondline[pos] == None:
                self.titledict[pos] = firstline[pos]
            if firstline[pos] != None and secondline[pos] != None:
                self.titledict[pos] = firstline[pos] + secondline[pos]
        for k,v in self.titledict.items():
            print(k, v)

    # form the new title of all nonnull cells of the title            
    def newtitlerow(self,  ):
        self.newtitle = []
        index = [0, 2, 4, 6, 10, 12, 13, 15, 17, 18, 19, 20]            
        for n in index:
            self.newtitle.append(self.titledict[n])
        logging.info(self.newtitle)

    # copy the rows that include students' marks according the index  
    def newrows(self,  ):
        self.newallrow = []
        self.newallrow.append(self.newtitle)
        for row in self.allrow:
            if re.compile(r'\d{12}').search(str(row[2])):
                newrow = []
                index = [0, 2, 4, 6, 10, 12, 13, 15, 17, 18, 19, 20]            
                for n in index:
                    newrow.append(row[n])
                self.newallrow.append(newrow)
        # logging.info(self.newallrow)     # it is too much letters to fit display

    # form the dictionary of the students, the value is a list of course, name and marks
    def studentsmarks(self,  ):
        self.newtitled = []
        index = [0, 2, 4, 6, 10, 12, 13, 15, 17, 18, 19, 20]            
        for n in index:
            self.newtitled.append(self.titledict[n])
        self.studentcoursesm.setdefault(self.titledict[2], self.newtitled)
        for row in self.allrow:
            if re.compile(r'\d{12}').search(str(row[2])):
                newrowd = []
                index = [0, 2, 4, 6, 10, 12, 13, 15, 17, 18, 19, 20]           
                for n in index:
                    newrowd.append(row[n])
                student = str(row[2])
                course = row[0]
                name = row[4]
                self.studentcoursesm.setdefault(student,[])
                self.studentcoursesm[student].append(newrowd)
        # logging.info(self.studentcoursesm)     # it is too much letters to fit display

    # output the workbook of all marks
    def writerows(self, dstdir):
        wb = openpyxl.Workbook()            
        sheet = wb.get_active_sheet()
        # Write title
        for col in range(len(self.newtitle)):
            sheet.cell(row = 2, column = col + 2).value = self.newtitle[col]
        # Write marks
        for row in range(len(self.newallrow)):     # row is begin from 0 to max
            for col in range(len(self.newallrow[row])):
                sheet.cell(row = row + 3, column = col + 2).value = str(self.newallrow[row][col])
        logging.info(str(row-1) + ' ' + str(col-1) + ': ' + str(self.newallrow[row-1][col-1]))
        newfullname = dstdir + '\\allmarks' + '.xlsx'
        wb.save(newfullname)
        print('All marks have been written.')

    # output the workbooks of marks of classes     
    def classmarks(self, dstdir):        
        pass
        
    # output the workbooks of marks of courses

    # output the workbooks of marks of students
    def studentmarks(self, studentnum, dstdir):
        wb = openpyxl.Workbook()            
        sheet = wb.get_active_sheet()
        # Write title
        for col in range(len(self.newtitle)):
            sheet.cell(row = 2, column = col + 2).value = self.newtitle[col]
        # Write marks
        for row in range(len(self.studentcoursesm[studentnum])):     # row is begin from 0 to max
            for col in range(len(self.studentcoursesm[studentnum][row])):
                sheet.cell(row = row + 3, column = col + 2).value = str(self.studentcoursesm[studentnum][row][col])
                logging.info(str(row) + str(col) + ': ' + str(self.studentcoursesm[studentnum][row][col]))
        newfullname = dstdir + '\\学号' + studentnum + '.xlsx'
        wb.save(newfullname)
        print('Marks of %s have been written.' % (studentnum))
        
    # print the marks of courses of the student
    def pstudentmarks(self, studentnum):
        print(self.newtitle)
        print(self.studentcoursesm[studentnum])
        # or
        STUDREG = re.compile(studentnum)
        for rowmark in self.newallrow:
            if STUDREG.search(str(rowmark[2])):
                self.studentm.append(rowmark)                
                print (rowmark)

#############################################
    def copyall(self, srcdir):
        twoline = 0
        self.classmark = {}    # for store marks of all courses, class as key, marks as the values
        self.allrow = []   # for store all row marks
        for t in self.classlist:
            self.coursemark = []   # for store marks of one course
            classname = t[0]
            file = t[1]            
            fullname = srcdir + '\\' + file
            logging.info(fullname)
            # copy marks ,add course name            
            wb = openpyxl.load_workbook( fullname)
            sheet = wb.get_active_sheet()
            for row in range(1,sheet.max_row):
                # the first element of every row is file name(include course name)
                rowmark = [file]               # for store marks of one row
                for col in range(1,sheet.max_column):
                    rowmark.append(sheet.cell(row = row, column = col).value)
                self.allrow.append(rowmark)      # add one row marks to form all row marks                
                self.coursemark.append(rowmark)      # add one row marks to form a course marks
                # get the title line. it has two lines.
##                print(rowmark[2])                
##                print(twoline)
##                input('any key')
                while(twoline == 1):
                    self.titlerow.append(rowmark)
                    twoline += 1
                if rowmark[2] == '学    号' and twoline == 0:
##                    print('in')
                    twoline += 1
                    self.titlerow.append(rowmark)
##            print(self.titlerow)                
#            logging.info(self.allrow)
#            input('any key')
            self.classmark.setdefault(classname, [])
            self.classmark[classname].append(self.coursemark)

    #  output one class' marks in one sheet. add the course name .     
    def writeclass(self, dstdir):        
        for classname,mark in self.classmark.items():
            wb = openpyxl.Workbook()            
            sheet = wb.get_active_sheet()
            # Write marks 
            row = 0
            logging.info(len(mark))
            logging.info(len(mark[0]))
            logging.info(len(mark[0][0]))
            for k in range(len(mark)):     # k is begin from 0 to max
                for r in range(len(mark[k])):
                    row += 1
                    for col in range(len(mark[k][r])):
                        sheet.cell(row = row, column = col + 2).value = mark[k][r][col]
            newfullname = dstdir + '\\cj-total-' + classname + '.xlsx'
            wb.save(newfullname)
            print('Marks of %s have been written.' % (classname))
#            input('any key') 
    

def main():
    SRCDIR = 'd:\\_PythonWorks\\excelOperate\\cj-2016201701'
    DSTDIR = 'd:\\_PythonWorks\\excelOperate\\cjcl-2016201701'
    CLASSRE = '-(\d{7}z?)\.' # CLASSREG = re.compile(r'\d{7}[zZ]?')
    
    collmark = collectmarks()
    CLASSLIST = collmark.findclass(SRCDIR,CLASSRE)
    input('finish findclass')
    collmark.findtitle(SRCDIR)
    input('finish findtitle')
    collmark.copyrows(SRCDIR)
    input('finish copyrows')
    collmark.findindex()
    input('finish findindex')
    collmark.newtitlerow()
    input('finish newtitlerow')
    collmark.newrows()
    input('finish newrows')
    collmark.studentsmarks()
    input('finish studentsmarks')
    collmark.writerows(DSTDIR)
    input('finish writerows')
    collmark.classmarks(DSTDIR)
    input('finish classmark')
    STUDNUM = input("please input student's number: ")
    collmark.studentmarks(STUDNUM,DSTDIR)
    input('finish studentmarks')
    collmark.pstudentmarks(STUDNUM)
    input('finish pstudentmarks')
#####################################     
    collmark.copyall(SRCDIR)
    input('finish copyall')
    collmark.writeclass(DSTDIR)
    input('finish writeclass')

if __name__ == '__main__':
    logging.critical('--------Start of program---------')
    main()
    logging.critical('--------End---------')

  

##    # copy all marks in one sheet. add the course name .    
##    def allmark(self, srcdir, dstdir):
##        self.studentm = []        
##        for t in self.classlist:
##            logging.info(t);logging.info(t[0]);logging.info(t[1])
##            classN = t[0]            
##            file = t[1]            
##            fullname = srcdir + '\\' + file
##            logging.info(fullname)
##            # copy marks ,add course name            
##            wb = openpyxl.load_workbook( fullname)
##            sheet = wb.get_active_sheet()
##            for row in range(1,sheet.max_row):
##                rowmark = []
##                rowmark.append(file)
##                for col in range(1,sheet.max_column):
##                    rowmark.append(sheet.cell(row = row, column = col).value)
##                self.studentm.append(rowmark)
##                logging.info(rowmark)
###                input('for debug')
##            logging.info(self.studentm)
##        # Write marks    
##        wb = openpyxl.Workbook()
##        sheet = wb.get_active_sheet()
##        for row in range(1,len(self.studentm) + 1):
##            for col in range(1,len(self.studentm[0]) + 1):
##                sheet.cell(row = row, column = col + 1).value = self.studentm[row][col]        
##        newfullname = dstdir + '\\cj-total' + '.xlsx'
##        wb.save(newfullname)   

    


