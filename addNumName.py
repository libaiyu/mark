#! python 3
# _*_ coding: utf_8  _*_
''' Read the number and name of the students from the file that include "学生名单".
Then write the number and name of the students to the new files that will note the mark for students.
find a bug. list of students need initial in the loop. not out of the loop. 2017-2-26.
Use getdir to get the directory.  next need to distict 1620604 and 1620604-5.  2017-3-22.
'''

import openpyxl
import os
import re
import logging

import getdir       # 2017-3-22

# logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.basicConfig( level = logging.ERROR, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.critical('--------Start of program---------')

regex = re.compile('(\w+)-(\d+)(-\d)?|(\w+)-([a-z]+)-(\d+)(-\d)?')   #  使用分组匹配
'''
>>> s = '模拟电子技术-performance-1523701-test'
>>> import re;regex = re.compile('(\w+)-(\d+)(-\d)?|(\w+)-([a-z]+)-(\d+)(-\d)?')
>>> regex.match( s)
<_sre.SRE_Match object; span=(0, 26), match='模拟电子技术-performance-1523701'>
>>> regex.match( s).groups()
(None, None, None, '模拟电子技术', 'performance', '1523701', None)
>>> import re
>>> s = '学生名单-1620604-5'
>>> regex = re.compile('(\w+)-(\d+)(-\d)?|(\w+)-([a-z]+)-(\d+)(-\d)?')
>>> regex.match( s)
<_sre.SRE_Match object; span=(0, 14), match='学生名单-1620604-5'>
>>> regex.match( s).groups()
('学生名单', '1620604', '-5', None, None, None, None)
'''

### find the file that include '学生名单' in file
##ChineseReg = re.compile(r'学生名单')
### find the class
##classReg = re.compile(r'\d{7}(-\d)?')
### find the course
##courseReg = re.compile(r'-([a-z]{3,11})-')

dirname = getdir.getdir()       # 2017-3-22

stud_dict = {}


matchlist = []       # List for storing the match file.
for file in os.listdir(dirname):
    print( file)
    # input('debug')
    if regex.match( file) is not None:      # match the first case. (\w+)-(\d+)(-\d)?
        matchgroup = regex.match( file).groups()
        matchlist.append( file)      #  get the match file list.
    if matchgroup[0]:    #  find the file that have "学生名单" in its name. 
        logging.info( matchgroup[0])    #  显示 "学生名单"
        
        # read the number and name of the students from the file that include "学生名单"
        fullname = dirname + '\\' + file
          
        if matchgroup[2]:
            className = matchgroup[1] + matchgroup[2]
        else:
            className = matchgroup[1]
        logging.info( className)        # find the class name in file name.
        students = [
            [ ],
            [ ],
            ['总分',],
            ['初始分',],
            ]

        wb = openpyxl.load_workbook(fullname)
        sheet = wb.get_active_sheet()
        # Read the name and number, then add to the student list.
        for row in range(1,sheet.max_row + 1):
            logging.debug(sheet['b'+str(row)].value)           #  学号在B列
            if sheet['b'+str(row)].value:                      #  学号在B列
                students[0].append(str(sheet['B'+str(row)].value))    #  学号在B列 
                students[1].append(sheet['d'+str(row)].value)         #  姓名在D列
                logging.debug(sheet['d'+str(row)].value)              #  姓名在d列
                students[2].append(65)         #  总分初始值为65
                students[3].append(65)         #  初始分设定为65
        # prepare dict for a new className
        stud_dict[ className] = students

# write the number and name of the students to the new files that will note the mark for students.
count = 0
for file in matchlist:
    matchgrp = regex.match( file).groups()
    if matchgrp[6] is not None:    # get the class name in the file that include course.
        className = matchgrp[5] + matchgrp[6]
    elif matchgrp[5] is not None:
        className = matchgrp[5]
    else:
        className = None
    logging.info( className)
    if className is not None:
        fullname = dirname + '\\' + file
        wb = openpyxl.Workbook()
        sheet = wb.get_active_sheet()
        # Write name and number.
        col = 2       
        for val in stud_dict[ className]:
            logging.info( val)
            for n in range( len( stud_dict[ className][0])):         
                logging.info( val[n])
                sheet.cell( row = 2 + n,column = col).value = val[n]
            col +=1
        wb.save( fullname)
        count += 1
        print('File for class: %s have been written! Total %d files.' % ( className, count))

logging.critical('-------End--------')




