# _*_ coding: utf_8  _*_
# Add the tag. initial value is 0.

import openpyxl
import os
import re
import logging

# logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.basicConfig( level = logging.ERROR, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.critical('--------Start of program---------')

courseReg = re.compile(r'-(\w{3,11})-')

performanceTag = [
##    ['学号', ],
##    ['姓名', ],
##    ['总分',],['初始分',],
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['提出问题',],
    ['回答问题',],
    ['作业1',],['作业2',],['作业3',],['作业4',],['作业5',],['作业6',],['作业7',],['作业8',],
    ]
labTag = [
##    ['学号', ],
##    ['姓名', ],
##    ['总分',],['初始分',],
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],['操作5',],['操作6',],['操作7',],['操作8',],
    ['数据1',],['数据2',],['数据3',],['数据4',],['数据5',],['数据6',],['数据7',],['数据8',],
    ['报告1',],['报告2',],['报告3',],['报告4',],    
    ]
designTag = [
##    ['学号', ],
##    ['姓名', ],
##    ['总分',],['初始分',],
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['设计1',],['设计2',],['设计3',],['设计4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],
    ['报告1',],['报告2',],    
    ]
practiceTag = [
##    ['学号', ],
##    ['姓名', ],
##    ['总分',],['初始分',],
    ['旷课',],
    ['迟到',],
    ['早退',],
    ['操作1',],['操作2',],['操作3',],['操作4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],
    ['报告1',],['报告2',],    
    ]
tagdict = {'performance':performance_tag,
           'lab':lab_tag,
           'design':design_tag,
           'practice':practice_tag}

dirname = 'd:\\_PythonWorks\\excelOperate\\pscj161702'
# open files one by one
count = 0
for file in os.listdir(dirname):
    if courseReg.search(file):
        courseType = courseReg.search(file)
        logging.info(courseType.group())
        fullname = dirname + '\\' + file
        if courseType.group(1) == 'performance':
            logging.info(courseType.group(1))            
            wb = openpyxl.load_workbook(fullname)
            sheet = wb.get_active_sheet()
            # Write the tag 
            col = 6
            for val in performanceTag:
                logging.info(val)
                sheet.cell(row = 2,column = col).value = val[0]
                for k in range(3,sheet.max_row + 1):
                    sheet.cell(row = k,column = col).value = 0
                col +=1
            
            wb.save(fullname)
            count += 1
            print('file: %s has been written!' % (file))

        if courseType.group(1) == 'lab':
            logging.info(courseType.group())            
            wb = openpyxl.load_workbook(fullname)
            sheet = wb.get_active_sheet()
            # Write the tag 
            col = 6
            for val in labTag:
                logging.info(val)
                logging.info(val[0])
                sheet.cell(row = 2 ,column = col).value = val[0]
                for k in range(3,sheet.max_row + 1):
                    sheet.cell(row = k,column = col).value = 0
                col +=1            
            wb.save(fullname)
            count += 1            
            print('file: %s has been written!' % (file))

        if courseType.group(1) == 'design':
            logging.info(courseType.group())            
            wb = openpyxl.load_workbook(fullname)
            sheet = wb.get_active_sheet()
            # Write the tag 
            col = 6
            for val in designTag:
                logging.info(val)
                sheet.cell(row = 2 ,column = col).value = val[0]
                for k in range(3,sheet.max_row + 1):
                    sheet.cell(row = k,column = col).value = 0
                col +=1            
            wb.save(fullname)
            count += 1
            print('file: %s has been written!' % (file))

        if courseType.group(1) == 'practice':
            logging.info(courseType.group())            
            wb = openpyxl.load_workbook(fullname)
            sheet = wb.get_active_sheet()
            # Write the tag 
            col = 6
            for val in practiceTag:
                logging.info(val)
                sheet.cell(row = 2 ,column = col).value = val[0]
                for k in range(3,sheet.max_row + 1):
                    sheet.cell(row = k,column = col).value = 0
                col +=1            
            wb.save(fullname)
            count += 1
            print('file: %s has been written!' % (file))
        
print('total %d files have finished Tag' % (count))

logging.critical('-------End--------')


