# _*_ coding: utf_8  _*_
 
import openpyxl
import os
import re
import logging


logging.basicConfig( level = logging.ERROR, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.critical('--------Start of program---------')

courseReg = re.compile(r'-(\w{3,11})-')

performanceTag = [
    ['学号', ],
    ['姓名', ],['初始分', ],
    ['旷课1',],['旷课2',],['旷课3',],
    ['提出问题1',],['提出问题2',],['提出问题3',],['提出问题4',],['提出问题5',],
    ['作业1',],['作业2',],['作业3',],['作业4',],['作业5',],['作业6',],['作业7',],['作业8',],
    ['回答问题1',],['回答问题2',],['回答问题3',],['回答问题4',],['回答问题5',],
    ['迟到1',],['迟到2',],['迟到3',],['迟到4',],['迟到5',],
    ['早退1',],['早退2',],['早退3',],['早退4',],['早退5',],
    ]
labTag = [
    ['学号', ],
    ['姓名', ],['初始分', ],
    ['旷课1',],['旷课2',],['旷课3',],
    ['操作1',],['操作2',],['操作3',],['操作4',],['操作5',],['操作6',],['操作7',],['操作8',],
    ['数据1',],['数据2',],['数据3',],['数据4',],['数据5',],['数据6',],['数据7',],['数据8',],
    ['报告1',],['报告2',],['报告3',],['报告4',],    
    ['迟到1',],['迟到2',],['迟到3',],['迟到4',],['迟到5',],
    ['早退1',],['早退2',],['早退3',],['早退4',],['早退5',],
    ]
designTag = [
    ['学号', ],
    ['姓名', ],['初始分', ],
    ['旷课1',],['旷课2',],['旷课3',],
    ['设计1',],['设计2',],['设计3',],['设计4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],
    ['报告1',],['报告2',],    
    ['迟到1',],['迟到2',],['迟到3',],['迟到4',],['迟到5',],
    ['早退1',],['早退2',],['早退3',],['早退4',],['早退5',],
    ]
practiceTag = [
    ['学号', ],
    ['姓名', ],['初始分', ],
    ['旷课1',],['旷课2',],['旷课3',],
    ['操作1',],['操作2',],['操作3',],['操作4',],
    ['数据1',],['数据2',],['数据3',],['数据4',],
    ['报告1',],['报告2',],    
    ['迟到1',],['迟到2',],['迟到3',],['迟到4',],['迟到5',],
    ['早退1',],['早退2',],['早退3',],['早退4',],['早退5',],
    ]

# open files one by one
for fileName in os.listdir('d:\\_PythonWorks\\execlOperate\\pscj161702'):
    if courseReg.search(fileName):
        courseType = courseReg.search(fileName)
        logging.info(courseType.group())
        
        if courseType.group(1) == 'performance':
            logging.info(courseType.group())            
            wb = openpyxl.load_workbook(fileName)
            sheet = wb.get_active_sheet()
            # Write 
            col = 2
            for val in performanceTag:
                logging.info(val)
                sheet.cell(row = 2 ,column = col).value = val[0]
                col +=1            
            wb.save(fileName)
            print('one file is written!')

        if courseType.group(1) == 'lab':
            logging.info(courseType.group())            
            wb = openpyxl.load_workbook(fileName)
            sheet = wb.get_active_sheet()
            # Write 
            col = 2
            for val in labTag:
                logging.info(val)
                sheet.cell(row = 2 ,column = col).value = val[0]
                col +=1            
            wb.save(fileName)
            print('one file is written!')

        if courseType.group(1) == 'design':
            logging.info(courseType.group())            
            wb = openpyxl.load_workbook(fileName)
            sheet = wb.get_active_sheet()
            # Write 
            col = 2
            for val in designTag:
                logging.info(val)
                sheet.cell(row = 2 ,column = col).value = val[0]
                col +=1            
            wb.save(fileName)
            print('one file is written!')

        if courseType.group(1) == 'practice':
            logging.info(courseType.group())            
            wb = openpyxl.load_workbook(fileName)
            sheet = wb.get_active_sheet()
            # Write 
            col = 2
            for val in practiceTag:
                logging.info(val)
                sheet.cell(row = 2 ,column = col).value = val[0]
                col +=1            
            wb.save(fileName)
            print('one file is written!')
        
print('finish Tag')

logging.critical('-------End--------')





