#! python3
# _*_ coding: utf_8  _*_
# copy marks in many workbooks and put them togather for every student.

import openpyxl
import os
import re
import logging

# logging.basicConfig( level = logging.ERROR, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )

class collectmarks():
    """    collect marks of the students in several courses.    
    """

    def __init__(self):
        
        self.filelist = []
        self.fulllist = []
        
        # the list for storing the classes and courses
        self.classlist = []
        # the list of tuple that incluse coursename+classname and filename
        self.courseclass = []
        # the row as the tag 
        self.tagrow = []
        # the list of all rows of one workbook. 
        self.bookrow = []
        # the list of the tag and its index
        self.taglist = []
        # after deleting null cell, we get the new tag. 
        
    def listfile(self, srcdir):
        # list the files in the given directory. produce the list of the file names
        k = 0
        self.filelist = []
        self.fulllist = []
        for file in os.listdir(srcdir):
            fullname = srcdir + '\\' + file
            self.filelist.append(file)
            self.fulllist.append(fullname)     
            print(k,file)
            k += 1
        return self.fulllist

    def findtag(self, workbook):
        # find the tag in given workbook
        twoline = 0            
        wb = openpyxl.load_workbook(workbook)
        sheet = wb.get_active_sheet()
        for row in range(1,sheet.max_row + 1):
            # rowmark: store marks of one row
            # the first element of every row is file name
            rowmark = [workbook]       
            for col in range(1,sheet.max_column + 1):
                rowmark.append(sheet.cell(row = row, column = col).value)
            # get the tag line. it has two lines.
            while(twoline == 1):
                self.tagrow.append(rowmark)
                twoline += 1
            if rowmark[2] == '学    号' and twoline == 0:
                twoline += 1
                self.tagrow.append(rowmark)
        print(self.tagrow)
 
    def findindex(self):
        # find the position of nonnull cells in tag,
        firstline = ['课程名', '序号', '学    号', None, '姓  名', None, '班  级', None, None, None,
                     '课堂', None,    '课堂',      '课堂',   None,  '课堂',   None, '实践成绩', '实验成绩',
                     '总成绩', '特殊原因', None, '录入状态', '备  注', None]
        secondline = ['课程名', None,    None,      None,    None,  None,   None,   None, None, None,
                      '平时成绩', None,'期中成绩',  '期末成绩', None, '总成绩', None,   None,      None,
                      None,     None,     None,    None,      None,   None]
        self.taglist = []
        self.taglist.append((0, firstline[0]))
        for pos in range(1,len(firstline)):
#            print(pos, firstline[pos], secondline[pos])
            if firstline[pos] != None and secondline[pos] == None:
                self.taglist.append((pos, firstline[pos]))
            elif firstline[pos] != None and secondline[pos] != None:
                self.taglist.append((pos, firstline[pos] + secondline[pos]))
            elif firstline[pos] == None and secondline[pos] != None:
                self.taglist.append((pos, secondline[pos]))
            else:
                self.taglist.append((pos, None))
                    
        for k in range(len(self.taglist)):
            print(self.taglist[k])

    def copybook(self, srcfile, dstdir):
        # copy every row in one workbooks to form a list
        self.bookrow = []   # for store a book row marks        
        # copy row marks ,add course name            
        wb = openpyxl.load_workbook( workbook)
        sheet = wb.get_active_sheet()
        for row in range(1,sheet.max_row + 1):
            # rowmark: store marks of one row
            # the first element of every row is file name
            rowmark = [file]       
            for col in range(1,sheet.max_column + 1):
                rowmark.append(sheet.cell(row = row, column = col).value)
            # add every row marks to form a book row marks
            self.bookrow.append(rowmark)
        newworkbook = dstdir + '\\' + 'copy-' + file
        wb.save(newworkbook)        

    def addremark(self, file, dstdir):
        # find the mark that failure to pass the exam in one workbook

        pass


def main():
    SRCDIR = 'd:\\_PythonWorks\\excelOperate\\cj-2016201701'
    DSTDIR = 'd:\\_PythonWorks\\excelOperate\\cjcl-2016201701'
    CLASSRE = '-(\d{7}z?)\.' # CLASSREG = re.compile(r'\d{7}[zZ]?')
    
    collmark = collectmarks()

    files = collmark.listfile(SRCDIR)
    input('finish listfile')
    coursenum = int(input('\n please input a number for 课程: '))
    book = files[coursenum]
    input('finish find book')

    
    collmark.findtag(book)
    input('finish findtag')
    collmark.findindex()
    input('finish findindex')


    collmark.copybook(book, DSTDIR)
    input('finish copybook')
    collmark.addremark(book, DSTDIR)
    input('finish addremark')
    
##    collmark.findbook(SRCDIR, '模拟电子技术基础', '1520603')
##    input('finish findbook')
##    CLASSLIST = collmark.findclass(SRCDIR,CLASSRE)
##    input('finish findclass')

if __name__ == '__main__':
    logging.critical('--------Start of program---------')
    main()
    logging.critical('--------End---------')


"""
    def findbook(self, srcdir, coursename, classname):
        # find the correct workbook by given course and class.
        for book in os.listdir(srcdir):
            COURSEREG = re.compile(coursename)
            CLASSREG = re.compile(classname)
            if COURSEREG.search(book) and CLASSREG.search(book):
                fullbook = srcdir + "\\" + book
                return fullbook

    def coursebook(self, srcdir, coursename):
        # find the correct workbook by course.
        coursebooks = []
        for book in os.listdir(srcdir):
            COURSEREG = re.compile(coursename)
            if COURSEREG.search(book):
                fullbook = srcdir + "\\" + book
                coursebooks.append(fullbook)
                return coursebooks

    def classbook(self, srcdir, classname):
        # find the correct workbook by course.
        classbooks = []
        for book in os.listdir(srcdir):
            CLASSREG = re.compile(classname)
            if CLASSREG.search(book):
                fullbook = srcdir + "\\" + book
                classbooks.append(fullbook)
                return classbooks
"""
