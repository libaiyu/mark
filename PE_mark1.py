#! python 3
# _*_ coding: utf_8  _*_

'''

'''

import openpyxl

import logging


MAXROW = 106

China_str = ['100米','跳远','铅球','８００米','男子成绩','女子成绩']
PE_items = ['100米','跳远','铅球','８００米']
TYPE = ['男子成绩','女子成绩']


measure_v = {}

def get_params():
    '''先获取所需的信息：男女，项目，测量值。

'''

    # TYPE[0] = '男子成绩' TYPE[1] = '女子成绩'
    gender = int( input('请选择。\n 0：男子成绩,1：女子成绩--'))
    # PE_items[0] = '100米', PE_items[1] = '跳远','铅球','８００米'
    item_name =  int( input('请选择。\n 0：100米,1:跳远,2:铅球,3:８００米--'))
    # '测得的成绩' 是字典里的键
    measure_val = float( input( '测得的成绩：'))
    return gender, item_name, measure_val

def get_mark( gender, item_name, measure_val):
    '''然后根据参数中提供的信息找到相应的成绩。
'''
    
    if gender:
        mark = mark_woman[PE_items[item_name]][measure_val]
    else:
        mark = mark_man[PE_items[item_name]][measure_val]
    print(mark)


# logging.basicConfig( level = logging.DEBUG, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.basicConfig( level = logging.ERROR, format = ' %(asctime)s - %(levelname)s - %(message)s' )
logging.critical('--------Start of program---------')



item_head = {}
mark_man = {}
mark_woman = {}

wb_val = {}


'''读成绩对照表，生成字典。以供后续查询。
'''
wb = openpyxl.load_workbook('高考体育成绩对照表.xlsx')
sheet = wb.get_active_sheet()
for col in range(1,sheet.max_column+1):
    key1 = col
    col_dict = {}
    for row in range(1,sheet.max_row+1):
        key2 = row
        val = sheet.cell( row=row, column=col).value # 给定的列与行的值。
##        print( type( val))
##        input()
##        if val is not None:
##            col_dict[ key2] = val
        col_dict[ key2] = val   #  字典－－遍历所有行。　｛行：值｝。
    wb_val[ key1] = col_dict    #  字典－－遍历所有列。　｛列：｛行：值｝｝。
print( key1, key2, val)

for col, row_val in wb_val.items():
    for row, val in row_val.items():
        if val in China_str:
            if val in PE_items:   #  如果值是项目名。
                item_head[val] = ( col, row)   #  字典--项目名为键，所在的（列，行）为值。
##print(item_head)

for val, ( col, row) in item_head.items():
    key1 = val    #  项目名
    item_man = {}
    item_woman = {}
    for r in range( row+2, MAXROW):
        key_m = wb_val[col][r]        #  男　的键　为　测量值
        val_m = wb_val[col+1][r]      #  男　的值　为　成绩
        item_man[key_m] = val_m       #  男　的字典　　测量值：成绩
        key_wo = wb_val[col+2][r]     #  女　的键　为　测量值
        val_wo = wb_val[col+3][r]     #  女　的值　为　成绩
        item_woman[key_wo] = val_wo   #  女　的字典
    mark_man[key1] = item_man         #  男　的字典的字典　项目名　为　键，（测量值：成绩）　为　值。
    mark_woman[key1] = item_woman     #  女　的字典的字典　项目名　为　键，（测量值：成绩）　为　值。
##print('男子成绩\n',mark_man,'\n')
##print('女子成绩\n',mark_woman,'\n')
##    input('any key, just for debug')




while True:

    gender, item_name, measure_val = get_params()

    get_mark( gender, item_name, measure_val)

    go_on = input('是否继续查成绩？\n其他:继续，"n":退出。')
    if go_on == 'n':
        break
    

logging.critical('-------End--------')






